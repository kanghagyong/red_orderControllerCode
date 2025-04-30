import pyautogui as pyautogui
from flask import Flask, send_from_directory, render_template_string
import os
import time
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
# import requests
from selenium.webdriver.common.alert import Alert
from typing import Optional
from selenium.webdriver.common.by import By
# 웹드라이버 생성
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.common.exceptions import WebDriverException, UnexpectedAlertPresentException, NoAlertPresentException, \
    TimeoutException
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
now = datetime.now()

pd.set_option('display.width', 320)
pd.set_option('display.max_columns', 20)
import tempfile

app3 = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
LIST_FOLDER = 'list'
DATA_FOLDER = 'data'
REDATA_FOLDER = 'reData'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app3.route('/')
def index():
    xlsx_files = [f for f in os.listdir(DATA_FOLDER) if f.endswith('.xlsx')]
    file_links = ''.join(f'<li><a href="/download/{f}">{f}</a></li>' for f in xlsx_files)
    first_fileupload =  render_template_string('''
        <div style="width:30%;float:left;">
            <h1>옵션 파일업</h1>
            <form action="/upload" method="post" enctype="multipart/form-data">
                <input type="file" name="file"><br><br>
                <input type="submit" value="파일생성">
            </form>
            <h1>다운로드 목록</h1>
            <ul>
                {{ files|safe }}
            </ul>
        </div>
    ''', files=file_links)

    xlsx_files1 = [f for f in os.listdir(REDATA_FOLDER) if f.endswith('.txt')]
    file_links1 = ''.join(f'<li><a href="/download2/{f}">{f}</a></li>' for f in xlsx_files1)
    first_fileupload1 =  render_template_string('''
        <div style="width:30%;float:left;">
            <h1>코드생성 파일업</h1>
            <form action="/upload2" method="post" enctype="multipart/form-data">
                <input type="file" name="file"><br><br>
                <input type="submit" value="파일생성">
            </form>
            <h1>다운로드 목록</h1>
            <ul>
                {{ files1|safe }}
            </ul>
        </div>
    ''', files1=file_links1)

    xlsx_files2 = [f for f in os.listdir(REDATA_FOLDER) if f.endswith('.txt')]
    file_links2 = ''.join(f'<li><a href="/download3/{f}">{f}</a></li>' for f in xlsx_files2)
    first_fileupload2 =  render_template_string('''
        <div style="width:30%;float:left;">
            <h1>코드생성 파일업(판)</h1>
            <form action="/upload3" method="post" enctype="multipart/form-data">
                <input type="file" name="file"><br><br>
                <input type="submit" value="파일생성">
            </form>
            <h1>다운로드 목록</h1>
            <ul>
                {{ files2|safe }}
            </ul>
        </div>
    ''', files2=file_links2)

    return first_fileupload+first_fileupload1+first_fileupload2

@app3.route('/upload', methods=['POST'])
def upload_file():
    from flask import request

    if 'file' not in request.files:
        return '파일이 없습니다.'
    file = request.files['file']
    if file.filename == '':
        return '파일 이름이 없습니다.'

    file.save(os.path.join(UPLOAD_FOLDER, file.filename))

    start_filename = UPLOAD_FOLDER+'/'+file.filename
    df_item_config = pd.read_excel(start_filename, sheet_name=0, engine='openpyxl', skiprows=1)
    uploadfile_option_check(df_item_config)
    time.sleep(3)
    return '파일이 업로드되었습니다!<br><a href="/">목록</a>'

@app3.route('/upload2', methods=['POST'])
def upload2_file():
    from flask import request

    if 'file' not in request.files:
        return '파일이 없습니다.'
    file = request.files['file']
    if file.filename == '':
        return '파일 이름이 없습니다.'

    file.save(os.path.join(UPLOAD_FOLDER, file.filename))

    start_filename = UPLOAD_FOLDER+'/'+file.filename
    df_item_config = pd.read_excel(start_filename, sheet_name=0, engine='openpyxl', nrows=1)
    df_item = pd.read_excel(start_filename, sheet_name=0, engine='openpyxl', skiprows=7)

    uploadfile_ordernum_creating(df_item_config, df_item)
    time.sleep(3)

    return '파일이 업로드되었습니다!<br><a href="/">목록</a>'

@app3.route('/upload3', methods=['POST'])
def upload3_file():
    from flask import request

    if 'file' not in request.files:
        return '파일이 없습니다.'
    file = request.files['file']
    if file.filename == '':
        return '파일 이름이 없습니다.'

    file.save(os.path.join(UPLOAD_FOLDER, file.filename))

    start_filename = UPLOAD_FOLDER+'/'+file.filename
    df_item_config = pd.read_excel(start_filename, sheet_name=0, engine='openpyxl', nrows=1)
    df_item = pd.read_excel(start_filename, sheet_name=0, engine='openpyxl', skiprows=7)

    uploadfile_ordernum_creating_pan(df_item_config, df_item)
    time.sleep(3)

    return '파일이 업로드되었습니다!<br><a href="/">목록</a>'

@app3.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

@app3.route('/download/<filename>')
def download(filename):
    # temp_dir = tempfile.gettempdir()
    return send_from_directory(DATA_FOLDER, filename, as_attachment=True)

@app3.route('/download2/<filename>')
def download2(filename):
    # temp_dir = tempfile.gettempdir()
    return send_from_directory(REDATA_FOLDER, filename, as_attachment=True)

@app3.route('/download3/<filename>')
def download3(filename):
    return send_from_directory(REDATA_FOLDER, filename, as_attachment=True)

# 옵션정리 및 코드생성용 파일 생성로직
def uploadfile_option_check(df_item_config):
    itemList_stationery = ['GSSBMTL', 'GSSBACM', 'GSSBSTP']
    itemList_koieditor = ['PHSTPAN']
    itemList_hk = ['GSKYHOT']
    itemList_tk = ['TPTKDFT', 'PRCAFIL']
    itemList_tee = ['CLTMMTS', 'CLTMHDS', 'CLTMSHS']
    itemList_memo = ['TPBLMEO1', 'TPBLMEO2', 'TPBLMEO3', 'TPBLMEO4']
    itemList_case = ['GSCAGBP', 'GSCAGBM','GSCAGBR','GSCAGBH','GSCASOP','GSCATPG','GSCAEPB','GSCACDP']
    totalList = len(df_item_config)
    nowtime = str(now.year) + '' + str(now.month) + '' + str(now.day) + '_' + str(now.hour) + '' + str(
        now.minute) + '' + str(now.second) + '' + str(now.microsecond)
    for item in range(totalList):
        # 상품코드 가져오기
        itemCode = df_item_config.iloc[item, 0]
        if itemCode in itemList_stationery:
            print('문방구 상품은 옵션 리스트를 따로 만들수 없습니다.')
            continue
        elif itemCode in itemList_koieditor:
            sampleOption_filename = LIST_FOLDER+'/koi_option_sample.xlsx'
            df_item = pd.read_excel(sampleOption_filename, sheet_name=0, engine='openpyxl', skiprows=1)
            if "," in df_item_config.iloc[item, 1]:
                papers_wgtlist = df_item_config.iloc[item, 1].split(",")
            else:
                papers_wgtlist = []
                papers_wgtlist.append(df_item_config.iloc[item, 1])
            if "," in df_item_config.iloc[item, 2]:
                resourceslist = df_item_config.iloc[item, 2].split(",")
            else:
                resourceslist = []
                resourceslist.append(df_item_config.iloc[item, 2])
            if df_item_config.iloc[item, 3] == 'x':
                afterpcs01_list = ['x']
            else:
                afterpcs01_list = df_item_config.iloc[item, 3].split(",")

            papers = []  # 용지리스트
            for pw in papers_wgtlist:
                pwdata = pw.split("_")
                if pwdata[0] in papers:
                    time.sleep(0.2)
                else:
                    papers.append(pwdata)

            i = 0
            for p in papers:
                for r in resourceslist:
                    for ap1 in afterpcs01_list:
                        df_item.loc[i, 'ItemCode'] = itemCode
                        df_item.loc[i, 'UrlLink'] = "https://www.redprinting.co.kr/ko/product/item/PH/PHSTPAN/detail/"+r
                        df_item.loc[i, 'Papers'] = p[0]
                        df_item.loc[i, 'WgtCod'] = p[1]
                        df_item.loc[i, 'Amount'] = 1
                        df_item.loc[i, 'AfterPcs01'] = ap1
                        df_item.loc[i, 'OrderCode'] = ""
                        df_item.loc[i, 'Price'] = ""
                        i = i + 1

            # print(i)
            new_filename = DATA_FOLDER+'/' + itemCode + '_' + nowtime + '.xlsx'
            df_item.to_excel(new_filename, sheet_name=itemCode, index=False, startrow=7)
            wb = load_workbook(new_filename)
            ws = wb.active
            ws['A1'] = 'UrlLink'
            ws['B1'] = 'ItemCode'
            ws['C1'] = 'ID'
            ws['D1'] = 'PW'

            ws['A2'] = ''
            ws['B2'] = itemCode
            ws['C2'] = 'x'
            ws['D2'] = 'x'

            wb.save(new_filename)
        elif itemCode in itemList_hk:
            sampleOption_filename = LIST_FOLDER + '/option_hk.xlsx'
            df_item = pd.read_excel(sampleOption_filename, sheet_name=0, engine='openpyxl', skiprows=1)
            if "," in df_item_config.iloc[item, 1]:
                color = df_item_config.iloc[item, 1].split(",")
            else:
                color = []
                color.append(df_item_config.iloc[item, 1])
            if "," in df_item_config.iloc[item, 2]:
                gori = df_item_config.iloc[item, 2].split(",")
            else:
                gori = []
                gori.append(df_item_config.iloc[item, 2])

            i = 0
            for c in color:
                for g in gori:
                    df_item.loc[i, 'ItemCode'] = itemCode
                    df_item.loc[i, 'Color'] = c
                    df_item.loc[i, 'Gori'] = g
                    df_item.loc[i, 'OrderCode'] = ""
                    df_item.loc[i, 'Price'] = ""
                    i = i + 1

            print(i)
            new_filename = DATA_FOLDER + '/' + itemCode + '_' + nowtime + '.xlsx'
            df_item.to_excel(new_filename, sheet_name=itemCode, index=False, startrow=7)
            wb = load_workbook(new_filename)
            ws = wb.active
            ws['A1'] = 'UrlLink'
            ws['B1'] = 'ItemCode'
            ws['C1'] = 'ID'
            ws['D1'] = 'PW'

            ws['A2'] = 'https://www.redprinting.co.kr/ko/product/item/' + itemCode[:2] + '/' + itemCode
            ws['B2'] = itemCode
            ws['C2'] = 'x'
            ws['D2'] = 'x'

            wb.save(new_filename)
        elif itemCode in itemList_tee:
            sampleOption_filename = LIST_FOLDER + '/option_tee.xlsx'
            df_item = pd.read_excel(sampleOption_filename, sheet_name=0, engine='openpyxl', skiprows=1)
            if "," in df_item_config.iloc[item, 1]:
                sizeop = df_item_config.iloc[item, 1].split(",")
            else:
                sizeop = []
                sizeop.append(df_item_config.iloc[item, 1])

            if "," in df_item_config.iloc[item, 2]:
                color = df_item_config.iloc[item, 2].split(",")
            else:
                color = []
                color.append(df_item_config.iloc[item, 2])

            # 사이즈 성인용 아동용 구분
            size = df_item_config.iloc[item, 3].split('-')
            sizeA = size[0]
            sizeC = size[1]
            if "," in df_item_config.iloc[item, 4]:
                printarea = df_item_config.iloc[item, 4].split(",")
            else:
                printarea = []
                printarea.append(df_item_config.iloc[item, 4])

            i = 0
            for so in sizeop:
                for c in color:
                    if so == 'A':
                        if "," in sizeA:
                            real_size = sizeA.split(",")
                        else:
                            real_size = []
                            real_size.append(sizeA)
                    elif so == 'C':
                        if "," in sizeC:
                            real_size = sizeC.split(",")
                        else:
                            real_size = []
                            real_size.append(sizeC)
                    for rs in real_size:
                        if rs != 'A':
                            for pa in printarea:
                                df_item.loc[i, 'ItemCode'] = itemCode
                                df_item.loc[i, 'SizeOp'] = so
                                df_item.loc[i, 'Color'] = c
                                df_item.loc[i, 'Size'] = rs
                                df_item.loc[i, 'PrintArea'] = pa
                                df_item.loc[i, 'OrderCode'] = ""
                                df_item.loc[i, 'Price'] = ""
                                i = i + 1

            print(i)
            new_filename = DATA_FOLDER + '/' + itemCode + '_' + nowtime + '.xlsx'
            df_item.to_excel(new_filename, sheet_name=itemCode, index=False, startrow=7)
            wb = load_workbook(new_filename)
            ws = wb.active
            ws['A1'] = 'UrlLink'
            ws['B1'] = 'ItemCode'
            ws['C1'] = 'ID'
            ws['D1'] = 'PW'

            ws['A2'] = 'https://www.redprinting.co.kr/ko/product/item/' + itemCode[:2] + '/' + itemCode
            ws['B2'] = itemCode
            ws['C2'] = 'x'
            ws['D2'] = 'x'

            wb.save(new_filename)
        elif itemCode in itemList_case:
            sampleOption_filename = LIST_FOLDER + '/option_case.xlsx'
            df_item = pd.read_excel(sampleOption_filename, sheet_name=0, engine='openpyxl', skiprows=1)
            if "," in df_item_config.iloc[item, 1]:
                pcstype = df_item_config.iloc[item, 1].split(",")
            else:
                pcstype = []
                pcstype.append(df_item_config.iloc[item, 1])

            if "," in df_item_config.iloc[item, 2]:
                company = df_item_config.iloc[item, 2].split(",")
            else:
                company = []
                company.append(df_item_config.iloc[item, 2])

            if "," in df_item_config.iloc[item, 3]:
                model = df_item_config.iloc[item, 3].split(",")
            else:
                model = []
                model.append(df_item_config.iloc[item, 3])
            i = 0
            for pcs in pcstype:
                for c in company:
                    for m in model:
                        df_item.loc[i, 'ItemCode'] = itemCode
                        df_item.loc[i, 'Type'] = pcs
                        df_item.loc[i, 'Brand'] = c
                        df_item.loc[i, 'Model'] = m
                        df_item.loc[i, 'OrderCode'] = ""
                        df_item.loc[i, 'Price'] = ""
                        i = i + 1

            print(i)
            new_filename = DATA_FOLDER + '/' + itemCode + '_' + str(item) + '_' + nowtime + '.xlsx'
            df_item.to_excel(new_filename, sheet_name=itemCode, index=False, startrow=7)
            wb = load_workbook(new_filename)
            ws = wb.active
            ws['A1'] = 'UrlLink'
            ws['B1'] = 'ItemCode'
            ws['C1'] = 'ID'
            ws['D1'] = 'PW'

            ws['A2'] = 'https://www.redprinting.co.kr/ko/product/item/' + itemCode[:2] + '/' + itemCode
            ws['B2'] = itemCode
            ws['C2'] = 'x'
            ws['D2'] = 'x'

            wb.save(new_filename)
        elif itemCode in itemList_memo:
            sampleOption_filename = LIST_FOLDER+'/option_sample2.xlsx'
            df_item = pd.read_excel(sampleOption_filename, sheet_name=0, engine='openpyxl', skiprows=1)
            if "," in df_item_config.iloc[item, 1]:
                papers_wgtlist = df_item_config.iloc[item, 1].split(",")
            else:
                papers_wgtlist = []
                papers_wgtlist.append(df_item_config.iloc[item, 1])
            if "," in df_item_config.iloc[item, 2]:
                dosulist = df_item_config.iloc[item, 2].split(",")
            else:
                dosulist = []
                dosulist.append(df_item_config.iloc[item, 2])
            if "," in df_item_config.iloc[item, 3]:
                sizelist = df_item_config.iloc[item, 3].split(",")
            else:
                sizelist = []
                sizelist.append(df_item_config.iloc[item, 3])
            if "," in str(df_item_config.iloc[item, 4]):
                amountlist = df_item_config.iloc[item, 4].split(",")
            else:
                amountlist = []
                amountlist.append(str(df_item_config.iloc[item, 4]))
            if df_item_config.iloc[item, 5] == 'x':
                afterpcs01_list = ['x']
            else:
                afterpcs01_list = df_item_config.iloc[item, 5].split(",")

            if df_item_config.iloc[item, 6] == 'x':
                afterpcs02_list = ['x']
            else:
                afterpcs02_list = df_item_config.iloc[item, 6].split(",")

            # print(papers_wgtlist)
            # print(dosulist)
            # print(sizelist)
            # print(amountlist)
            # print(afterpcs01_list)
            # print(afterpcs02_list)
            # print(afterpcs03_list)
            # print(afterpcs04_list)
            # print(afterpcs05_list)

            papers = []  # 용지리스트
            wgt = []  # 용지무게 리스트
            for pw in papers_wgtlist:
                pwdata = pw.split("_")
                if pwdata[0] in papers:
                    time.sleep(0.2)
                else:
                    papers.append(pwdata)

            i = 0
            for p in papers:
                for d in dosulist:
                    for s in sizelist:
                        for ap1 in afterpcs01_list:
                            for ap2 in afterpcs02_list:
                                for a in amountlist:
                                    df_item.loc[i, 'ItemCode'] = 'TPBLMEO'
                                    df_item.loc[i, 'Papers'] = p[0]
                                    df_item.loc[i, 'WgtCod'] = p[1]
                                    df_item.loc[i, 'Dosu'] = d
                                    df_item.loc[i, 'Sizes'] = s
                                    df_item.loc[i, 'Amount'] = a
                                    df_item.loc[i, 'AfterPcs01'] = ap1
                                    df_item.loc[i, 'AfterPcs02'] = ap2
                                    df_item.loc[i, 'OrderCode'] = ""
                                    df_item.loc[i, 'Price'] = ""
                                    i = i + 1

            print(i)
            new_filename = DATA_FOLDER+'/' + itemCode + '_' + nowtime + '.xlsx'
            df_item.to_excel(new_filename, sheet_name=itemCode, index=False, startrow=7)
            wb = load_workbook(new_filename)
            ws = wb.active
            ws['A1'] = 'UrlLink'
            ws['B1'] = 'ItemCode'
            ws['C1'] = 'ID'
            ws['D1'] = 'PW'

            ws['A2'] = 'https://www.redprinting.co.kr/ko/product/item/'+itemCode[:2]+'/TPBLMEO'
            ws['B2'] = 'TPBLMEO'
            ws['C2'] = 'x'
            ws['D2'] = 'x'

            wb.save(new_filename)
        elif itemCode in itemList_tk:
            sampleOption_filename = LIST_FOLDER+'/option_tk.xlsx'
            df_item = pd.read_excel(sampleOption_filename, sheet_name=0, engine='openpyxl', skiprows=1)
            if "," in df_item_config.iloc[item, 1]:
                papers_wgtlist = df_item_config.iloc[item, 1].split(",")
            else:
                papers_wgtlist = []
                papers_wgtlist.append(df_item_config.iloc[item, 1])
            if "," in df_item_config.iloc[item, 2]:
                dosulist = df_item_config.iloc[item, 2].split(",")
            else:
                dosulist = []
                dosulist.append(df_item_config.iloc[item, 2])
            if "," in df_item_config.iloc[item, 3]:
                sizelist = df_item_config.iloc[item, 3].split(",")
            else:
                sizelist = []
                sizelist.append(df_item_config.iloc[item, 3])
            if "," in str(df_item_config.iloc[item, 4]):
                amountlist = df_item_config.iloc[item, 4].split(",")
            else:
                amountlist = []
                amountlist.append(str(df_item_config.iloc[item, 4]))
            if df_item_config.iloc[item, 5] == 'x':
                afterpcs01_list = ['x']
            else:
                afterpcs01_list = df_item_config.iloc[item, 5].split(",")

            papers = []  # 용지리스트
            wgt = []  # 용지무게 리스트
            for pw in papers_wgtlist:
                pwdata = pw.split("_")
                if pwdata[0] in papers:
                    time.sleep(0.2)
                else:
                    papers.append(pwdata)

            i = 0
            for p in papers:
                for d in dosulist:
                    for s in sizelist:
                        for ap1 in afterpcs01_list:
                            for a in amountlist:
                                df_item.loc[i, 'ItemCode'] = itemCode
                                df_item.loc[i, 'Papers'] = p[0]
                                df_item.loc[i, 'WgtCod'] = p[1]
                                df_item.loc[i, 'Dosu'] = d
                                df_item.loc[i, 'Sizes'] = s
                                df_item.loc[i, 'Amount'] = a
                                df_item.loc[i, 'APcs01'] = ap1
                                df_item.loc[i, 'OrderCode'] = ""
                                df_item.loc[i, 'Price'] = ""
                                i = i + 1

            print(i)
            new_filename = DATA_FOLDER+'/' + itemCode + '_' + nowtime + '.xlsx'
            df_item.to_excel(new_filename, sheet_name=itemCode, index=False, startrow=7)
            wb = load_workbook(new_filename)
            ws = wb.active
            ws['A1'] = 'UrlLink'
            ws['B1'] = 'ItemCode'
            ws['C1'] = 'ID'
            ws['D1'] = 'PW'

            ws['A2'] = 'https://www.redprinting.co.kr/ko/product/item/'+itemCode[:2]+'/'+itemCode
            ws['B2'] = itemCode
            ws['C2'] = 'x'
            ws['D2'] = 'x'

            wb.save(new_filename)
        else:
            sampleOption_filename = LIST_FOLDER+'/option_sample2.xlsx'
            df_item = pd.read_excel(sampleOption_filename, sheet_name=0, engine='openpyxl', skiprows=1)
            if "," in df_item_config.iloc[item, 1]:
                papers_wgtlist = df_item_config.iloc[item, 1].split(",")
            else:
                papers_wgtlist = []
                papers_wgtlist.append(df_item_config.iloc[item, 1])
            if "," in df_item_config.iloc[item, 2]:
                dosulist = df_item_config.iloc[item, 2].split(",")
            else:
                dosulist = []
                dosulist.append(df_item_config.iloc[item, 2])
            if "," in df_item_config.iloc[item, 3]:
                sizelist = df_item_config.iloc[item, 3].split(",")
            else:
                sizelist = []
                sizelist.append(df_item_config.iloc[item, 3])
            if "," in str(df_item_config.iloc[item, 4]):
                amountlist = df_item_config.iloc[item, 4].split(",")
            else:
                amountlist = []
                amountlist.append(str(df_item_config.iloc[item, 4]))
            if df_item_config.iloc[item, 5] == 'x':
                afterpcs01_list = ['x']
            else:
                afterpcs01_list = df_item_config.iloc[item, 5].split(",")

            if df_item_config.iloc[item, 6] == 'x':
                afterpcs02_list = ['x']
            else:
                afterpcs02_list = df_item_config.iloc[item, 6].split(",")

            if df_item_config.iloc[item, 7] == 'x':
                afterpcs03_list = ['x']
            else:
                afterpcs03_list = df_item_config.iloc[item, 7].split(",")

            if df_item_config.iloc[item, 8] == 'x':
                afterpcs04_list = ['x']
            else:
                afterpcs04_list = df_item_config.iloc[item, 8].split(",")

            if df_item_config.iloc[item, 9] == 'x':
                afterpcs05_list = ['x']
            else:
                afterpcs05_list = df_item_config.iloc[item, 9].split(",")

            # print(papers_wgtlist)
            # print(dosulist)
            # print(sizelist)
            # print(amountlist)
            # print(afterpcs01_list)
            # print(afterpcs02_list)
            # print(afterpcs03_list)
            # print(afterpcs04_list)
            # print(afterpcs05_list)

            papers = []  # 용지리스트
            wgt = []  # 용지무게 리스트
            for pw in papers_wgtlist:
                pwdata = pw.split("_")
                if pwdata[0] in papers:
                    time.sleep(0.2)
                else:
                    papers.append(pwdata)

            i = 0
            for p in papers:
                for d in dosulist:
                    for s in sizelist:
                        for ap1 in afterpcs01_list:
                            for a in amountlist:
                                # if ap1 == '유광':
                                for ap2 in afterpcs02_list:
                                    for ap3 in afterpcs03_list:
                                        for ap4 in afterpcs04_list:
                                            for ap5 in afterpcs05_list:
                                                df_item.loc[i, 'ItemCode'] = itemCode
                                                df_item.loc[i, 'Papers'] = p[0]
                                                df_item.loc[i, 'WgtCod'] = p[1]
                                                df_item.loc[i, 'Dosu'] = d
                                                df_item.loc[i, 'Sizes'] = s
                                                df_item.loc[i, 'Amount'] = a
                                                df_item.loc[i, 'AfterPcs01'] = ap1
                                                df_item.loc[i, 'AfterPcs02'] = ap2
                                                df_item.loc[i, 'AfterPcs03'] = ap3
                                                df_item.loc[i, 'AfterPcs04'] = ap4
                                                df_item.loc[i, 'AfterPcs05'] = ap5
                                                df_item.loc[i, 'OrderCode'] = ""
                                                df_item.loc[i, 'Price'] = ""
                                                i = i + 1

            print(i)
            new_filename = DATA_FOLDER+'/' + itemCode + '_' + nowtime + '.xlsx'
            df_item.to_excel(new_filename, sheet_name=itemCode, index=False, startrow=7)
            wb = load_workbook(new_filename)
            ws = wb.active
            ws['A1'] = 'UrlLink'
            ws['B1'] = 'ItemCode'
            ws['C1'] = 'ID'
            ws['D1'] = 'PW'

            ws['A2'] = 'https://www.redprinting.co.kr/ko/product/item/'+itemCode[:2]+'/'+itemCode
            ws['B2'] = itemCode
            ws['C2'] = 'x'
            ws['D2'] = 'x'

            wb.save(new_filename)

# 로그인 체크 프로세스
def login_check_proc(userid, userpw, itemUrl, driver, itemCode):
    time.sleep(1)
    driver.get(itemUrl)

    time.sleep(1)
    driver.find_element(By.CSS_SELECTOR, '#header > div.main-home-header-warp > div > div.header-right > ul > li.select-my > a').click()

    # 로그인화면 아이디 및 패스워드 입력
    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
    driver.find_element(By.ID, 'mb_id').click()
    driver.find_element(By.ID, 'mb_id').send_keys(userid)
    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
    driver.find_element(By.ID, 'mb_password').click()
    driver.find_element(By.ID, 'mb_password').send_keys(userpw)
    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
    driver.find_element(By.ID, 'btnLogin').click()
    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )

    # driver.execute_script('window.scrollTo(0, 200)')

# 일반적인 상품 주문관리 코드 생성 로직
def uploadfile_ordernum_creating(df_item_config, df_item):
    itemList_card = ['BCSPDFT', 'BNSTDFT', 'PRCAFIL', 'BNSTMAS']
    itemList_hk = ['GSKYHOT']
    itemList_sticker = ['STTHELP', 'STTHSQU', 'STTHUSR']
    itemList_sticker2 = ['STPADIY']
    itemList_stationery = ['GSSBMTL', 'GSSBACM', 'GSSBSTP']
    itemList_tee = ['CLTMMTS', 'CLTMHDS', 'CLTMSHS']
    itemList_case = ['GSCAGBP', 'GSCAGBM','GSCAGBR','GSCAGBH','GSCASOP','GSCATPG','GSCAEPB','GSCACDP']
    itemList_offset = ['NCDFDFT', 'NCDFQLT', 'NCDFCPN', 'NCCDDFT']
    itemList_memo = ['TPBLMEO']
    itemList_note = ['GSNTMIS']
    itemList_calendar = ['PRCLSTD']
    itemList_poster = ['PRPORSO', 'PRPOWTT', 'PRPOXSP', 'PRPOXPO', 'PRPOWHT', 'WBXXXXX']
    itemList_kring = ['GSSRCUT', 'GSSRPRT']
    itemList_pendunte = ['GSPASGC']
    itemList_PRCAXPO = ['PRCAXPO']
    itemList_LFXXXXX = ['LFXXXXX', 'PRPOXXX', 'PRLFXXX', 'STTHCIC', 'STCUXXX']
    # 현재 화면 크기 가져오기
    screen_width, screen_height = pyautogui.size()

    # 70% 크기로 계산
    window_width = int(screen_width * 0.7)
    window_height = int(screen_height * 0.7)

    options = ChromeOptions()
    options.add_argument('--blink-settings=imagesEnabled=false')
    options.set_capability('goog:loggingPrefs', {'browser': 'SEVERE'})

    driver = webdriver.Chrome(options=options)
    driver.implicitly_wait(3)
    totalList = len(df_item)

    driver.set_window_size(window_width, window_height)
    # (선택) 위치도 조절하고 싶다면
    driver.set_window_position(0, 0)

    # 상품코드 가져오기
    itemUrl = df_item_config.iloc[0,0]
    itemCode = df_item_config.iloc[0,1]
    userid = df_item_config.iloc[0,2]
    userpw = df_item_config.iloc[0,3]

    if userid == "x":
        userid = 'red_openmarket' #red_openmarket, #redprinting
    if userpw == "x":
        userpw = 'red4874#' #red4874# , #redprinting#1234
    login_check_proc(userid, userpw, itemUrl, driver, itemCode)

    # newFileData = "아이템코드\t용지\t용지무게\t인쇄도수\t가로\t세로\t사이즈\t수량\t코팅후가공\t타공후가공\t귀돌이후가공\t재단후가공\t아일렛후가공\t주문관리코드\t금액\n"
    try:
        # 각 상품코드 별 어떤 형태인지 체크
        if itemCode in itemList_card:
            for i in range(totalList):
                paper = df_item.iloc[i, 1]
                wgtcod = df_item.iloc[i, 2]
                dosu = df_item.iloc[i, 3]
                size = df_item.iloc[i, 4]
                amount = df_item.iloc[i, 5]
                apcs1 = df_item.iloc[i, 6]
                apcs2 = df_item.iloc[i, 7]
                apcs3 = df_item.iloc[i, 8]
                apcs4 = df_item.iloc[i, 9]
                apcs5 = df_item.iloc[i, 10]
                # sizesplit = size.split("*")
                # wid_size = sizesplit[0]
                # hei_size = sizesplit[1]
                # newFileData =newFileData +itemCode+"\t"+str(paper)+"\t"+str(wgtcod)+"\t"+str(dosu)+"\t"+str(wid_size)+"\t"+str(hei_size)+"\t"
                # newFileData =newFileData +str(size)+"\t"+str(amount)+"\t"+str(apcs1)+"\t"+str(apcs2)+"\t"+str(apcs3)+"\t"+str(apcs4)+"\t"+str(apcs5)

                driver.execute_script('window.scrollTo(0, 100)')
                time.sleep(0.5)
                paper_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "paperSelectBoxItText")) )
                if paper_text.text != paper:
                    # 상품 페이지 용지선택
                    time.sleep(0.5)
                    #paperSelectBoxItContainer
                    driver.find_element(By.ID, 'paperSelectBoxIt').click()
                    time.sleep(0.5)
                    driver.find_element(By.LINK_TEXT, paper).click()
                time.sleep(0.5)
                if wgtcod != "x":
                    # 상품 페이지 G수 선택
                    if driver.find_element(By.ID, 'paper_sub_selectSelectBoxItText').text != str(wgtcod):
                        #paper_sub_selectSelectBoxItContainer
                        driver.find_element(By.ID, 'paper_sub_selectSelectBoxIt').click()
                        time.sleep(1)
                        driver.find_element(By.LINK_TEXT, str(wgtcod)).click()
                time.sleep(0.5)
                if dosu != "":
                    # 상품 페이지 인쇄도수 선택
                    docu_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "soduSelectBoxItText")) )
                    if docu_text.text != dosu:
                        #soduSelectBoxItContainer
                        driver.find_element(By.ID, 'soduSelectBoxIt').click()
                        WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                        driver.find_element(By.LINK_TEXT, dosu).click()
                time.sleep(0.5)
                if itemCode == 'PRCAFIL':
                    time.sleep(0.2)
                else:
                    WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    size_text = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.ID, "sizeSelectBoxItText")))
                    if size_text.text != size:
                        # 사이즈 직접입력 선택 sizeSelectBoxItContainer
                        driver.find_element(By.ID, 'sizeSelectBoxIt').click()
                        WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                        driver.find_element(By.LINK_TEXT, size).click()
                        if size == '사이즈직접입력':
                            driver.find_element(By.ID, 'CUT_WDT').click()
                            driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.DELETE)
                            driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.DELETE)
                            driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.DELETE)
                            driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.DELETE)
                            driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                            driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                            driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                            driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                            time.sleep(0.5)
                            driver.find_element(By.ID, 'CUT_WDT').send_keys('280')

                            driver.find_element(By.ID, 'CUT_HGH').click()
                            driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.DELETE)
                            driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.DELETE)
                            driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.DELETE)
                            driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.DELETE)
                            driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                            driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                            driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                            driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                            time.sleep(0.5)
                            driver.find_element(By.ID, 'CUT_HGH').send_keys('400')

                driver.find_element(By.XPATH, '//*[@id="WRK_HGH"]').click()

                time.sleep(0.5)
                if amount != "":
                    driver.execute_script("productOrder.check_PRN_CNT();")
                    number1_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "number1")) )
                    if number1_text.get_attribute('value') != str(amount):
                        time.sleep(0.5)
                        driver.find_element(By.ID, 'number1').click()
                        driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                        driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                        driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                        driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                        driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                        time.sleep(0.5)
                        driver.find_element(By.ID, 'number1').send_keys(str(amount))
                    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )

                time.sleep(0.5)
                apcs_nowstatus = driver.find_element(By.ID, 'priceCalcResult').get_attribute('value')#priceCalcResult, opt_string
                if apcs1 == 'x':
                    time.sleep(0.5)
                else:
                    if apcs1 == "무광":
                        if 'COT_DFT' in apcs_nowstatus:
                            time.sleep(0.2)
                        else:
                            # 상품 페이지 코팅 선택
                            driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                        time.sleep(0.5)
                        # 상품 페이지 무광코팅 선택
                        driver.execute_script("productOrder.opt_select('COT_DFT','MA');")
                    elif apcs1 == "유광":
                        if 'COT_DFT' in apcs_nowstatus:
                            time.sleep(0.2)
                        else:
                            # 상품 페이지 코팅 선택
                            driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                        time.sleep(0.5)
                        # 상품 페이지 유광코팅 선택
                        driver.execute_script("productOrder.opt_select('COT_DFT','GL');")
                    else:
                        if 'COT_DFT' in apcs_nowstatus:
                            # 코팅 후가공 다시 선택 시 선택해제됨.
                            driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                        else:
                            time.sleep(0.2)

                if apcs2 == "x":
                    time.sleep(0.2)
                else:
                    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                    driver.execute_script("productOrder.opt_use_yn('HOL_DFT', '');")
                    aftpcs2 = apcs2.split('-')
                    apcs2_cnt = aftpcs2[0]
                    apcs2_val = aftpcs2[1]
                    driver.find_element(By.ID, 'HOL_DFT_IN1').click()
                    driver.find_element(By.ID, 'HOL_DFT_IN1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'HOL_DFT_IN1').send_keys(Keys.BACK_SPACE)
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'HOL_DFT_IN1').send_keys(str(apcs2_cnt))
                    # 라디오 버튼 네임값이 동일한 것 찾아서 apcs2 값 동일한 것으로 클릭 후 멈춤
                    afterPcs2List = driver.find_elements(By.NAME, 'HOL_DFT_MM')
                    for ap2l in afterPcs2List:
                        if ap2l.get_attribute('value') == apcs2_val:
                            ap2l.click()
                            break

                if apcs3 == 'x':
                    time.sleep(0.2)
                else:
                    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                    driver.execute_script("productOrder.opt_use_yn('ROU_DFT', '');")
                    aftpcs3 = apcs3.split('_')
                    apcs3_size = aftpcs3[0]
                    apcs3_location = str(aftpcs3[1])
                    afterPcs3List = driver.find_elements(By.NAME, 'ROU_DFT_MM')
                    for ap3l in afterPcs3List:
                        if ap3l.get_attribute('value') == apcs3_size:
                            ap3l.click()
                            break
                    apcs3location = driver.find_element(By.ID, 'ROU_DFT_ALL')

                    if apcs3_location == '1111' or apcs3_location == 'all':
                        time.sleep(0.2)
                    else:
                        if apcs3location.is_selected():
                            apcs3location.click()
                        else:
                            time.sleep(0.2)
                        if apcs3_location[0] == '1':
                            driver.execute_script("productOrder.rou_dft_select('DFXLT');")
                        if apcs3_location[1] == '1':
                            driver.execute_script("productOrder.rou_dft_select('DFXRT');")
                        if apcs3_location[2] == '1':
                            driver.execute_script("productOrder.rou_dft_select('DFXLB');")
                        if apcs3_location[3] == '1':
                            driver.execute_script("productOrder.rou_dft_select('DFXRB');")

                if apcs4 == 'x':
                    time.sleep(0.2)
                else:
                    WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    if apcs4 == '방풍커팅':
                        driver.execute_script("productOrder.opt_checked('CUT_ZUN', 'ZDWND');")

                if apcs5 == 'x':
                    time.sleep(0.2)
                else:
                    WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    if apcs5 == '은색아일렛':
                        time.sleep(0.2)
                        # driver.find_element(By.ID, 'ILT_DFT_IN1').click()
                    elif apcs5 == '구리색아일렛':
                        driver.find_element(By.XPATH, '//*[@id="ILT_DFT_SUB_DIV2"]/label[2]/div[1]/img')
                        # driver.find_element(By.ID, 'ILT_DFT_IN2').click()

                driver.find_element(By.XPATH, '//*[@id="subject"]')
                time.sleep(0.5)
                driver.execute_script('window.scrollTo(0, 1500)')
                time.sleep(0.5)
                total_price = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.ID, "PRICE_DIS2")))
                tprice = total_price.text
                time.sleep(0.5)
                # driver.find_element(By.ID, 'direct_order_btn').click()
                driver.execute_script("productOrder.order_validate('pot_create');")
                time.sleep(3)
                al = Alert(driver)
                al.accept()
                time.sleep(0.5)
                # imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
                # df_item.loc[i, 'OrderCode'] = imsiordernum #주문관리코드생성후추가
                # print(i, "번째 TOTAL_PRICE : ", total_price, "pot_tmp_cod : ", imsiordernum)
                print(i, "번째 TOTAL_PRICE : ", tprice)
                driver.execute_script('window.location.reload();')
                driver.execute_script('window.scrollTo(0, 100);')
                time.sleep(1)
        elif itemCode in itemList_sticker:
            for i in range(totalList):
                wait1 = WebDriverWait(driver, 5)
                wait2 = WebDriverWait(driver, 10)
                paper = df_item.iloc[i, 1]
                wgtcod = df_item.iloc[i, 2]
                dosu = df_item.iloc[i, 3]
                size = df_item.iloc[i, 4]
                amount = df_item.iloc[i, 5]
                apcs1 = df_item.iloc[i, 6]
                apcs2 = df_item.iloc[i, 7]
                apcs3 = df_item.iloc[i, 8]
                apcs4 = df_item.iloc[i, 9]
                apcs5 = df_item.iloc[i, 10]
                sizesplit = size.split("*")
                wid_size = sizesplit[0]
                hei_size = sizesplit[1]

                time.sleep(2)
                paper_text = wait1.until( EC.visibility_of_element_located((By.ID, "paperSelectBoxItText")) )
                if paper_text.text != paper:
                    # 상품 페이지 용지선택
                    wait2.until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                    # paperSelectBoxItContainer
                    driver.find_element(By.ID, 'paperSelectBoxIt').click()
                    wait2.until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                    driver.find_element(By.LINK_TEXT, paper).click()

                wait1.until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                if wgtcod != "":
                    # 상품 페이지 G수 선택
                    wgt_text = wait1.until( EC.visibility_of_element_located((By.ID, "paper_sub_selectSelectBoxItText")) )
                    if wgt_text.text != str(wgtcod):
                        #paper_sub_selectSelectBoxItContainer
                        driver.find_element(By.ID, 'paper_sub_selectSelectBoxIt').click()
                        wait2.until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                        driver.find_element(By.LINK_TEXT, str(wgtcod)).click()

                wait2.until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                if dosu != "":
                    # 상품 페이지 인쇄도수 선택
                    docu_text = wait1.until( EC.visibility_of_element_located((By.ID, "soduSelectBoxItText")) )
                    if docu_text.text != dosu:
                        #soduSelectBoxItContainer
                        driver.find_element(By.ID, 'soduSelectBoxIt').click()
                        wait2.until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                        driver.find_element(By.LINK_TEXT, dosu).click()

                driver.execute_script("productOrder.order_type_change('editor');")  # 코이에디터 선택
                time.sleep(2)

                wait2.until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                size_text = wait1.until( EC.visibility_of_element_located((By.ID, "sizeSelectBoxItText")) )
                if size_text.text != '사이즈직접입력':
                    # 사이즈 직접입력 선택 sizeSelectBoxItContainer
                    driver.find_element(By.ID, 'sizeSelectBoxIt').click()
                    wait2.until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                    driver.find_element(By.LINK_TEXT, '사이즈직접입력').click()

                CUT_WDT_text = wait1.until( EC.visibility_of_element_located((By.ID, "CUT_WDT")) )
                if CUT_WDT_text.get_attribute('value') != wid_size:
                    # 상품 페이지 사이즈 직접입력 사이즈 입력
                    driver.find_element(By.ID, 'CUT_WDT').click()
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'CUT_WDT').send_keys(wid_size)

                CUT_HGH_text = wait1.until( EC.visibility_of_element_located((By.ID, "CUT_HGH")) )
                if CUT_HGH_text.get_attribute('value') != hei_size:
                    driver.find_element(By.ID, 'CUT_HGH').click()
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'CUT_HGH').send_keys(hei_size)

                wait2.until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                driver.find_element(By.XPATH, '//*[@id="WRK_HGH"]').click()
                wait2.until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                driver.execute_script("productOrder.check_PRN_CNT();")
                number1_text = wait1.until( EC.visibility_of_element_located((By.ID, "number1")) )
                if number1_text.get_attribute('value') != str(amount):
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'number1').click()
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'number1').send_keys(str(amount))

                wait2.until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                apcs_nowstatus = driver.find_element(By.ID, 'opt_string').get_attribute('value')
                if apcs1 == "무광":
                    if 'COT_DFT' in apcs_nowstatus:
                        time.sleep(0.2)
                    else:
                        # 상품 페이지 코팅 선택
                        driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                    wait2.until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                    driver.execute_script("productOrder.opt_select('COT_DFT','MA');")# 상품 페이지 무광코팅 선택
                elif apcs1 == "유광":
                    if 'COT_DFT' in apcs_nowstatus:
                        time.sleep(0.2)
                    else:
                        # 상품 페이지 코팅 선택
                        driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                    wait2.until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                    driver.execute_script("productOrder.opt_select('COT_DFT','GL');")# 상품 페이지 유광코팅 선택
                else:
                    if 'COT_DFT' in apcs_nowstatus:
                        driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');") #코팅 후가공 다시 선택 시 선택해제됨.
                    else:
                        time.sleep(0.2)
                wait2.until( EC.invisibility_of_element_located((By.ID, 'overlay')) )

                if apcs4 == "묶음재단":
                    driver.execute_script("productOrder.opt_checked('CUT_DFT', 'DFXXX');")  # 상품 페이지 묶음재단 선택
                elif apcs4 == "개별재단":
                    driver.execute_script("productOrder.opt_checked('CUT_DFT', 'DFITM');")  # 상품 페이지 개별재단 선택
                else:
                    time.sleep(0.2)
                wait1.until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                driver.find_element(By.XPATH, '//*[@id="WRK_HGH"]').click()

                print(i, "번째 : ", paper, "|", wgtcod, "|", dosu, "|", size, "|", amount, "|", apcs1, "|", apcs4)
                # total_price = driver.find_element(By.ID, 'TOTAL_PRICE').text
                # df_item.loc[i, 'Price'] = tprice
                time.sleep(0.5)
                # try:
                #     driver.execute_script("productOrder.order_validate('pot_create');")
                #     time.sleep(1)
                #     driver.find_element(By.ID, 'direct_order_btn').click()
                #     try:
                #         # alert가 뜰 때까지 기다리기 (최대 3초)
                #         alert = wait1.until(EC.alert_is_present())
                #         alert.accept()
                #         wait1.until(EC.element_to_be_clickable((By.ID, 'direct_order_btn'))).click()
                #     except TimeoutException:
                #         wait1.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                #
                # except Exception as e:
                #     print("예외 발생:", e)

                print("생성완료.")
                driver.execute_script('window.scrollTo(0, 100);')
                time.sleep(2)
        elif itemCode in itemList_hk:
            # newFileData = "아이템코드\t색상\t열쇠고리\t주문관리코드\t금액\n"
            for i in range(totalList):
                color = df_item.iloc[i, 1]
                gori = df_item.iloc[i, 2]
                # newFileData =newFileData +itemCode+"\t"+str(color)+"\t"+str(gori)
                #print(newFileData)
                driver.execute_script('window.scrollTo(0, 100)')
                time.sleep(0.5)
                driver.execute_script("productOrder.check_GSKYHOT('color','"+color+"');")
                time.sleep(3)
                if gori == 'x':
                    time.sleep(1)
                else:
                    if gori[:2] == 'KR':
                        driver.execute_script("productOrder.opt_web_pcs_dtl_grp('SUB_MTR_KR', '');")
                        time.sleep(0.5)
                        driver.find_element(By.ID, 'SUB_MTR_KR_SUB_SELECTSelectBoxIt').click()
                        time.sleep(0.5)
                        parent_ele = driver.find_element(By.ID, 'SUB_MTR_KR_SUB_SELECTSelectBoxItOptions')
                        li_elements = parent_ele.find_elements(By.TAG_NAME, "li")
                        for li in li_elements:
                            data_val = li.get_attribute("data-val")
                            if data_val == gori:
                                li.click()
                                break
                    elif gori[:2] == 'BN':
                        driver.execute_script("productOrder.opt_web_pcs_dtl_grp('SUB_MTR_BN', '');")
                        time.sleep(0.5)
                        driver.find_element(By.ID, 'SUB_MTR_BN_SUB_SELECTSelectBoxIt').click()
                        time.sleep(0.5)
                        parent_ele = driver.find_element(By.ID, 'SUB_MTR_BN_SUB_SELECTSelectBoxItOptions')
                        li_elements = parent_ele.find_elements(By.TAG_NAME, "li")
                        for li in li_elements:
                            data_val = li.get_attribute("data-val")
                            if data_val == gori:
                                li.click()
                                break
                    else:
                        time.sleep(1)

                driver.execute_script('window.scrollTo(0, 400)')
                time.sleep(0.5)
                total_price = driver.find_element(By.ID, 'TOTAL_PRICE').text
                time.sleep(0.5)
                driver.find_element(By.ID, 'direct_order_btn').click()
                time.sleep(1)
                al = Alert(driver)
                al.accept()
                time.sleep(0.5)
                imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
                print(i, "번째 TOTAL_PRICE : ", total_price, "pot_tmp_cod : ", imsiordernum)
                # newFileData = newFileData+"\t"+imsiordernum+"\t"+total_price+"\n"
                # if (i+1) % 5 == 0:
                #     nowtime = str(now.year)+''+str(now.month)+''+str(now.day)+'_'+str(now.hour)+''+str(now.minute)
                #     new_filename = REDATA_FOLDER+'/'+userid+'_'+itemCode+"_"+str(i+1)+'_'+nowtime+'.txt'
                #     with open(new_filename, 'w', encoding='ansi') as file:
                        # file.write(newFileData)

                # driver.execute_script('window.location.reload();')
                time.sleep(2)
        elif itemCode in itemList_tee:
            # newFileData = "아이템코드\t사이즈옵션\t컬러\t사이즈\t인쇄영역\t주문관리코드\t금액\n"
            for i in range(totalList):
                sizeop = df_item.iloc[i, 1]
                color = df_item.iloc[i, 2]
                size = df_item.iloc[i, 3]
                if '-' in df_item.iloc[i, 4]:
                    paList = df_item.iloc[i, 4].split("-")
                else:
                    paList = []
                    paList.append(df_item.iloc[i, 4])
                # newFileData =newFileData +itemCode+"\t"+str(sizeop)+"\t"+str(color)+"\t"+str(size)+"\t"+df_item.iloc[i, 4]
                driver.execute_script('window.scrollTo(0, 100)')
                time.sleep(0.5)
                if sizeop == 'A':
                    time.sleep(0.5)
                else:
                    driver.execute_script("productOrder.check_CLST('', 'size_option', '"+sizeop+"');")
                    time.sleep(2)

                if size == 'B':
                    time.sleep(0.5)
                else:
                    driver.execute_script("productOrder.check_CLST('', 'size', '"+str(size)+"');")
                    time.sleep(2)

                if color == '53':
                    time.sleep(0.5)
                else:
                    driver.execute_script("productOrder.check_clothing('','clothing_color','"+str(color)+"');")
                    time.sleep(2)

                time.sleep(0.5)
                if len(paList) == 1:
                    if paList[0] == 'CL011':
                        time.sleep(0.5)
                    elif paList[0] == 'CL001':
                        driver.execute_script("productOrder.check_CLST('','clothing_area', 'CL001');")
                        time.sleep(1)
                    elif paList[0] == 'CL002':
                        driver.execute_script("productOrder.check_CLST('','clothing_area', 'CL002');")
                        time.sleep(1)
                        driver.execute_script("productOrder.check_CLST('','clothing_area', 'CL011');")
                        time.sleep(1)
                else:
                    for pa in paList:
                        if pa == 'CL011':
                            time.sleep(0.5)
                        elif pa == 'CL001':
                            driver.execute_script("productOrder.check_CLST('','clothing_area', 'CL001');")
                            time.sleep(1)
                        elif pa == 'CL002':
                            driver.execute_script("productOrder.check_CLST('','clothing_area', 'CL002');")
                            time.sleep(1)

                driver.execute_script("productOrder.order_type_change('editor');")  # 코이에디터 선택
                time.sleep(1)
                driver.execute_script('window.scrollTo(0, 500);')
                time.sleep(0.5)
                total_price = driver.find_element(By.ID, 'TOTAL_PRICE').text
                time.sleep(0.5)
                driver.find_element(By.ID, 'direct_order_btn').click()
                time.sleep(3)
                al = Alert(driver)
                al.accept()
                time.sleep(0.5)
                imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
                print(i, "번째 TOTAL_PRICE : ", total_price, "pot_tmp_cod : ", imsiordernum)
                # print(i, "번째 TOTAL_PRICE : ", total_price)
                # newFileData = newFileData+"\t"+imsiordernum+"\t"+total_price+"\n"
                # if (i+1) % 5 == 0:
                #     nowtime = str(now.year)+''+str(now.month)+''+str(now.day)+'_'+str(now.hour)+''+str(now.minute)
                #     new_filename = REDATA_FOLDER+'/backup/'+userid+'_'+itemCode+"_"+str(i+1)+'_'+nowtime+'.txt'
                #     with open(new_filename, 'w', encoding='ansi') as file:
                #         file.write(newFileData)

                # driver.execute_script('window.location.reload();')
                time.sleep(2)
        elif itemCode in itemList_case:
            # newFileData = "아이템코드\t자재종류\t종류\t기종\t주문관리코드\t금액\n"
            for i in range(totalList):
                pcsType = df_item.iloc[i, 1]
                brand = df_item.iloc[i, 2]
                model = df_item.iloc[i, 3]

                # newFileData =newFileData +itemCode+"\t"+str(pcsType)+"\t"+str(brand)+"\t"+str(model)
                driver.execute_script('window.scrollTo(0, 100)')
                time.sleep(0.5)
                if itemCode == 'GSCASOP':
                    if pcsType == '블랙':
                        time.sleep(0.5)
                    else:
                        if pcsType == '스카이블루':
                            driver.execute_script("$('#gsca_option2_1').click();")
                            time.sleep(2)
                        elif pcsType == '퍼플':
                            driver.execute_script("$('#gsca_option2_2').click();")
                            time.sleep(2)
                        elif pcsType == '핑크':
                            driver.execute_script("$('#gsca_option2_3').click();")
                            time.sleep(2)
                        elif pcsType == '옐로우':
                            driver.execute_script("$('#gsca_option2_4').click();")
                            time.sleep(2)
                        elif pcsType == '화이트':
                            driver.execute_script("$('#gsca_option2_5').click();")
                            time.sleep(2)
                else:
                    if pcsType == '유광':
                        time.sleep(0.5)
                    elif pcsType == '일반':
                        time.sleep(0.5)
                    else:
                        driver.execute_script("$('#gsca_option1_1').click();")
                        time.sleep(2)

                brand_ele = driver.find_element(By.ID, 'gsca_option3')
                select = Select(brand_ele)
                selectBrand = brand
                for bop in select.options:
                    if bop.text == selectBrand:
                        select.select_by_visible_text(selectBrand)

                time.sleep(2)
                model_ele = driver.find_element(By.ID, 'gsca_type')
                select2 = Select(model_ele)
                selectModel = model
                for mop in select2.options:
                    if mop.text == selectModel:
                        select2.select_by_visible_text(selectModel)
                time.sleep(1)

                driver.execute_script("productOrder.order_type_change('editor');")
                time.sleep(1)

                driver.execute_script('window.scrollTo(0, 300);')
                time.sleep(0.5)
                total_price = driver.find_element(By.ID, 'TOTAL_PRICE').text
                time.sleep(0.5)
                driver.find_element(By.ID, 'direct_order_btn').click()
                time.sleep(5)
                driver.find_element(By.ID, 'direct_order_btn').click()
                time.sleep(3)
                al = Alert(driver)
                al.accept()
                time.sleep(0.5)
                imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
                print(i, "번째 TOTAL_PRICE : ", total_price, "pot_tmp_cod : ", imsiordernum)
                # print(i, "번째 TOTAL_PRICE : ", total_price)
                # newFileData = newFileData+"\t"+imsiordernum+"\t"+total_price+"\n"
                # if (i+1) % 5 == 0:
                #     nowtime = str(now.year)+''+str(now.month)+''+str(now.day)+'_'+str(now.hour)+''+str(now.minute)
                #     new_filename = REDATA_FOLDER+'/backup/'+userid+'_'+itemCode+"_"+str(i+1)+'_'+nowtime+'.txt'
                #     with open(new_filename, 'w', encoding='ansi') as file:
                #         file.write(newFileData)

                # driver.execute_script('window.location.reload();')
                time.sleep(2)
        elif itemCode in itemList_sticker2:
            for i in range(totalList):
                paper = df_item.iloc[i, 1]
                wgtcod = df_item.iloc[i, 2]
                dosu = df_item.iloc[i, 3]
                size = df_item.iloc[i, 4]
                amount = df_item.iloc[i, 5]
                apcs1 = df_item.iloc[i, 6]
                apcs2 = df_item.iloc[i, 7]
                apcs3 = df_item.iloc[i, 8]
                apcs4 = df_item.iloc[i, 9]
                apcs5 = df_item.iloc[i, 10]
                time.sleep(2)
                paper_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "paperSelectBoxItText")) )
                if paper_text.text != paper:
                    # 상품 페이지 용지선택
                    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                    # paperSelectBoxItContainer
                    driver.find_element(By.ID, 'paperSelectBoxIt').click()
                    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                    driver.find_element(By.LINK_TEXT, paper).click()

                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                if wgtcod != "x":
                    # 상품 페이지 G수 선택
                    wgt_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "paper_sub_selectSelectBoxItText")) )
                    if wgt_text.text != str(wgtcod):
                        #paper_sub_selectSelectBoxItContainer
                        driver.find_element(By.ID, 'paper_sub_selectSelectBoxIt').click()
                        WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                        driver.find_element(By.LINK_TEXT, str(wgtcod)).click()

                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                if dosu != "":
                    # 상품 페이지 인쇄도수 선택
                    docu_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "soduSelectBoxItText")) )
                    if docu_text.text != dosu:
                        #soduSelectBoxItContainer
                        driver.find_element(By.ID, 'soduSelectBoxIt').click()
                        WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                        driver.find_element(By.LINK_TEXT, dosu).click()

                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                size_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "sizeSelectBoxItText")) )
                if size_text.text != size:
                    # 사이즈 직접입력 선택 sizeSelectBoxItContainer
                    driver.find_element(By.ID, 'sizeSelectBoxIt').click()
                    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                    driver.find_element(By.LINK_TEXT, size).click()


                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                driver.find_element(By.XPATH, '//*[@id="WRK_HGH"]').click()
                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                driver.execute_script("productOrder.check_PRN_CNT();")
                number1_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "number1")) )
                if number1_text.get_attribute('value') != str(amount):
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'number1').click()
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'number1').send_keys(str(amount))

                driver.execute_script("productOrder.order_type_change('editor');")  # 코이에디터 선택
                time.sleep(1)

                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                apcs_nowstatus = driver.find_element(By.ID, 'opt_string').get_attribute('value')
                if apcs1 == 'x':
                    time.sleep(0.5)
                else:
                    driver.execute_script(" productOrder.opt_use_yn('COT_DFT', 'SID_S'); ")  # 상품 페이지 무광코팅 선택
                    if apcs1 == "무광":
                        WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                        driver.execute_script("productOrder.opt_select('COT_DFT','MA');")# 상품 페이지 무광코팅 선택
                    elif apcs1 == "유광":
                        WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                        driver.execute_script("productOrder.opt_select('COT_DFT','GL');")# 상품 페이지 유광코팅 선택

                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )

                if apcs2 == "x":
                    time.sleep(0.5)
                else:
                    driver.execute_script("productOrder.opt_use_yn('PAK_POL', '');")  # 상품 페이지 무광코팅 선택
                    time.sleep(0.5)
                    driver.execute_script("productOrder.opt_checked('PAK_POL', '"+apcs2+"');")  # 상품 페이지 유광코팅 선택

                    time.sleep(0.5)
                WebDriverWait(driver, 5).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                driver.find_element(By.XPATH, '//*[@id="WRK_HGH"]').click()
                total_price = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "PRICE_DIS2")) )

                driver.execute_script("$('#chk_notice_confirm').click();")  # 확인했습니다.
                time.sleep(2)

                tprice = total_price.text
                # total_price = driver.find_element(By.ID, 'TOTAL_PRICE').text
                # df_item.loc[i, 'Price'] = tprice
                time.sleep(0.5)
                driver.find_element(By.ID, 'direct_order_btn').click()
                # time.sleep(3)
                # al = Alert(driver)
                # al.accept()
                time.sleep(0.5)
                imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
                print(i, "번째 TOTAL_PRICE : ", tprice, "pot_tmp_cod : ", imsiordernum)
                driver.execute_script('window.location.reload();')
                driver.execute_script('window.scrollTo(0, 100);')
                time.sleep(2)
        elif itemCode in itemList_note:
            for i in range(totalList):
                paper = df_item.iloc[i, 1]
                wgtcod = df_item.iloc[i, 2]
                dosu = df_item.iloc[i, 3]
                size = df_item.iloc[i, 4]
                amount = df_item.iloc[i, 5]
                apcs1 = df_item.iloc[i, 6]
                apcs2 = df_item.iloc[i, 7]
                apcs3 = df_item.iloc[i, 8]
                apcs4 = df_item.iloc[i, 9]
                apcs5 = df_item.iloc[i, 10]
                time.sleep(2)
                if size != 'SIZE_SUB_RADIO_1':
                    driver.execute_script('$("#SIZE_SUB_RADIO_2").click();');
                    time.sleep(0.5)
                    driver.execute_script('$("#SIZE_SUB_RADIO_2").click();');

                p_select = driver.find_element(By.ID, 'paper')
                pselect = Select(p_select)
                for poption in pselect.options:
                    if paper == poption.text:
                        pselect.select_by_visible_text(poption.text)
                time.sleep(1)

                if wgtcod != '':
                    time.sleep(0.5)

                driver.execute_script("productOrder.check_PRN_CNT();")
                number1_text = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.ID, "number1")))
                if number1_text.get_attribute('value') != str(amount):
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'number1').click()
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'number1').send_keys(str(amount))

                driver.find_element(By.XPATH, '//*[@id="INN_DFT_SUB_DIV"]/label/div[1]/img').click()
                driver.execute_script('window.scrollTo(0, 800);')
                apcs1_select = driver.find_element(By.ID, 'INN_DFT_SUB_SELECT')
                apcs1select = Select(apcs1_select)
                for apcs1option in apcs1select.options:
                    if apcs1 == apcs1option.text:
                        apcs1select.select_by_visible_text(apcs1option.text)
                time.sleep(1)

                if apcs2 != '':
                    driver.execute_script('$("#' + apcs2 + '").click();');
                    time.sleep(0.5)
                    driver.execute_script('$("#' + apcs2 + '").click();');

                WebDriverWait(driver, 5).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                total_price = WebDriverWait(driver, 5).until(
                    EC.visibility_of_element_located((By.ID, "PRICE_DIS2")))
                tprice = total_price.text
                time.sleep(0.5)

                # driver.execute_script("$('#chk_notice_confirm').click();")  # 확인했습니다.
                # time.sleep(2)

                driver.execute_script("productOrder.order_validate('pot_create');")  # 주문관리코드생성
                time.sleep(0.5)
                # driver.find_element(By.ID, 'direct_order_btn').click()
                imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
                print(i, "번째 TOTAL_PRICE : ", tprice, "pot_tmp_cod : ", imsiordernum)
                driver.execute_script('window.location.reload();')
                driver.execute_script('window.scrollTo(0, 100);')
                time.sleep(2)
        elif itemCode in itemList_calendar:
            for i in range(totalList):
                paper = df_item.iloc[i, 1]
                wgtcod = df_item.iloc[i, 2]
                dosu = df_item.iloc[i, 3]
                size = df_item.iloc[i, 4]
                amount = df_item.iloc[i, 5]
                apcs1 = df_item.iloc[i, 6]
                apcs2 = df_item.iloc[i, 7]
                apcs3 = df_item.iloc[i, 8]
                apcs4 = df_item.iloc[i, 9]
                apcs5 = df_item.iloc[i, 10]
                time.sleep(2)

                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                if dosu != "":
                    # 상품 페이지 인쇄도수 선택
                    docu_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "soduSelectBoxItText")) )
                    if docu_text.text != dosu:
                        #soduSelectBoxItContainer
                        driver.find_element(By.ID, 'soduSelectBoxIt').click()
                        WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                        driver.find_element(By.LINK_TEXT, dosu).click()

                # paper_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "paperSelectBoxItText")) )
                # papertext =paper_text.get_attribute("data-val")
                driver.find_element(By.ID, 'paperSelectBoxIt').click()
                if paper == 'RAU':
                    # 상품 페이지 용지선택
                    driver.find_element(By.XPATH, '//*[@id="paperSelectBoxItOptions"]/li[1]').click()
                    time.sleep(0.5)
                elif paper == 'DGP':
                    driver.find_element(By.XPATH, '//*[@id="paperSelectBoxItOptions"]/li[2]').click()
                    time.sleep(0.5)
                elif paper == 'DLS':
                    driver.find_element(By.XPATH, '//*[@id="paperSelectBoxItOptions"]/li[3]').click()
                    time.sleep(0.5)


                # WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                # if wgtcod != "x":
                #     # 상품 페이지 G수 선택
                #     wgt_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "paper_sub_selectSelectBoxItText")) )
                #     if wgt_text.text != str(wgtcod):
                #         #paper_sub_selectSelectBoxItContainer
                #         driver.find_element(By.ID, 'paper_sub_selectSelectBoxIt').click()
                #         WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                #         driver.find_element(By.LINK_TEXT, str(wgtcod)).click()
                #

                # WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                # size_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "sizeSelectBoxItText")) )
                # if size_text.text != size:
                #     # 사이즈 직접입력 선택 sizeSelectBoxItContainer
                #     driver.find_element(By.ID, 'sizeSelectBoxIt').click()
                #     WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                #     driver.find_element(By.LINK_TEXT, size).click()
                #
                #
                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                driver.find_element(By.XPATH, '//*[@id="WRK_HGH"]').click()
                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                driver.execute_script("productOrder.check_PRN_CNT();")
                number1_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "number1")) )
                if number1_text.get_attribute('value') != str(amount):
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'number1').click()
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'number1').send_keys(str(amount))

                # driver.execute_script("productOrder.order_type_change('editor');")  # 코이에디터 선택
                # time.sleep(1)

                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                # apcs_nowstatus = driver.find_element(By.ID, 'opt_string').get_attribute('value')
                if apcs1 == '블랙':
                    time.sleep(0.5)
                elif apcs1 == '진곤색':
                    driver.execute_script("$('#TRI_DFT_RADIO_G').click();")  # 상품 페이지 무광코팅 선택
                    time.sleep(0.5)
                    driver.execute_script("$('#TRI_DFT_RADIO_G').click();")  # 상품 페이지 무광코팅 선택
                    time.sleep(0.5)
                elif apcs1 == '아이보리':
                    driver.execute_script("$('#TRI_DFT_RADIO_I').click();")  # 상품 페이지 무광코팅 선택
                    time.sleep(0.5)
                    driver.execute_script("$('#TRI_DFT_RADIO_I').click();")  # 상품 페이지 무광코팅 선택
                    time.sleep(0.5)

                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                if apcs2 == '검정색':
                    time.sleep(0.5)
                elif apcs2 == '흰색':
                    driver.execute_script("$('#RIN_DFT_RADIO_1').click();")  # 상품 페이지 무광코팅 선택
                    time.sleep(0.5)
                    driver.execute_script("$('#RIN_DFT_RADIO_1').click();")  # 상품 페이지 무광코팅 선택
                    time.sleep(0.5)
                elif apcs2 == '금색':
                    driver.execute_script("$('#RIN_DFT_RADIO_2').click();")  # 상품 페이지 무광코팅 선택
                    time.sleep(0.5)
                    driver.execute_script("$('#RIN_DFT_RADIO_2').click();")  # 상품 페이지 무광코팅 선택
                    time.sleep(0.5)
                elif apcs2 == '은색':
                    driver.execute_script("$('#RIN_DFT_RADIO_3').click();")  # 상품 페이지 무광코팅 선택
                    time.sleep(0.5)
                    driver.execute_script("$('#RIN_DFT_RADIO_3').click();")  # 상품 페이지 무광코팅 선택
                    time.sleep(0.5)

                WebDriverWait(driver, 5).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                driver.find_element(By.XPATH, '//*[@id="WRK_HGH"]').click()
                total_price = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "PRICE_DIS2")) )
                tprice = total_price.text

                # driver.execute_script("$('#chk_notice_confirm').click();")  # 확인했습니다.
                # time.sleep(2)
                # total_price = driver.find_element(By.ID, 'TOTAL_PRICE').text
                # df_item.loc[i, 'Price'] = tprice
                time.sleep(0.5)
                driver.execute_script("productOrder.order_validate('pot_create');")  # 확인했습니다.
                # driver.find_element(By.ID, 'direct_order_btn').click()
                # time.sleep(3)
                # al = Alert(driver)
                # al.accept()
                time.sleep(0.5)
                # imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
                print(i, "번째 TOTAL_PRICE : ", tprice) #, "pot_tmp_cod : ", imsiordernum
                driver.execute_script('window.location.reload();')
                driver.execute_script('window.scrollTo(0, 100);')
                time.sleep(2)
        elif itemCode in itemList_poster:
            for i in range(totalList):
                paper = df_item.iloc[i, 1]
                wgtcod = df_item.iloc[i, 2]
                dosu = df_item.iloc[i, 3]
                size = df_item.iloc[i, 4]
                amount = df_item.iloc[i, 5]
                apcs1 = df_item.iloc[i, 6]
                apcs2 = df_item.iloc[i, 7]
                apcs3 = df_item.iloc[i, 8]
                apcs4 = df_item.iloc[i, 9]
                apcs5 = df_item.iloc[i, 10]
                time.sleep(2)
                paper_text = WebDriverWait(driver, 5).until(
                    EC.visibility_of_element_located((By.ID, "paperSelectBoxItText")))
                if paper_text.text != paper:
                    # 상품 페이지 용지선택
                    WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    # paperSelectBoxItContainer
                    driver.find_element(By.ID, 'paperSelectBoxIt').click()
                    WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    driver.find_element(By.LINK_TEXT, paper).click()

                WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                if wgtcod != "x":
                    # 상품 페이지 G수 선택
                    wgt_text = WebDriverWait(driver, 5).until(
                        EC.visibility_of_element_located((By.ID, "paper_sub_selectSelectBoxItText")))
                    if wgt_text.text != str(wgtcod):
                        # paper_sub_selectSelectBoxItContainer
                        driver.find_element(By.ID, 'paper_sub_selectSelectBoxIt').click()
                        WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                        driver.find_element(By.LINK_TEXT, str(wgtcod)).click()

                WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                if dosu != "":
                    # 상품 페이지 인쇄도수 선택
                    docu_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "soduSelectBoxItText")) )
                    if docu_text.text != dosu:
                        driver.find_element(By.ID, 'soduSelectBoxIt').click()
                        WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                        driver.find_element(By.LINK_TEXT, dosu).click()

                WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                size_text = WebDriverWait(driver, 5).until(
                    EC.visibility_of_element_located((By.ID, "sizeSelectBoxItText")))
                if size_text.text != size:
                    # 사이즈 직접입력 선택 sizeSelectBoxItContainer
                    driver.find_element(By.ID, 'sizeSelectBoxIt').click()
                    WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    driver.find_element(By.LINK_TEXT, size).click()
                    if size == '사이즈직접입력':
                        driver.find_element(By.ID, 'CUT_WDT').click()
                        driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.DELETE)
                        driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.DELETE)
                        driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.DELETE)
                        driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.DELETE)
                        driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                        time.sleep(0.5)
                        driver.find_element(By.ID, 'CUT_WDT').send_keys('280')

                        driver.find_element(By.ID, 'CUT_HGH').click()
                        driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.DELETE)
                        driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.DELETE)
                        driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.DELETE)
                        driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.DELETE)
                        driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                        time.sleep(0.5)
                        driver.find_element(By.ID, 'CUT_HGH').send_keys('400')

                driver.find_element(By.XPATH, '//*[@id="WRK_HGH"]').click()

                # time.sleep(1)
                # WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                # if dosu == 'PRT_MRD':
                #     color_btn = driver.find_element(By.XPATH, '//*[@id="tr_palette"]/td/div/label[1]/div[1]/img')
                #     time.sleep(0.5)
                # elif dosu == 'PRT_MOG':
                #     color_btn = driver.find_element(By.XPATH, '//*[@id="tr_palette"]/td/div/label[2]/div[1]/img')
                #     time.sleep(0.5)
                # elif dosu == 'PRT_MGN':
                #     color_btn = driver.find_element(By.XPATH, '//*[@id="tr_palette"]/td/div/label[3]/div[1]/img')
                #     time.sleep(0.5)
                # elif dosu == 'PRT_MBL':
                #     color_btn = driver.find_element(By.XPATH, '//*[@id="tr_palette"]/td/div/label[4]/div[1]/img')
                #     time.sleep(0.5)
                # elif dosu == 'PRT_MNV':
                #     color_btn = driver.find_element(By.XPATH, '//*[@id="tr_palette"]/td/div/label[5]/div[1]/img')
                #     time.sleep(0.5)
                # elif dosu == 'PRT_MFG':
                #     color_btn = driver.find_element(By.XPATH, '//*[@id="tr_palette"]/td/div/label[6]/div[1]/img')
                #     time.sleep(0.5)
                # elif dosu == 'PRT_MBK':
                #     color_btn = driver.find_element(By.XPATH, '//*[@id="tr_palette"]/td/div/label[7]/div[1]/img')
                #     time.sleep(0.5)
                # else:
                #     time.sleep(0.5)
                # color_btn.click()
                # time.sleep(0.5)

                WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                driver.find_element(By.XPATH, '//*[@id="WRK_HGH"]').click()
                WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                driver.execute_script("productOrder.check_PRN_CNT();")
                number1_text = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.ID, "number1")))
                if number1_text.get_attribute('value') != str(amount):
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'number1').click()
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'number1').send_keys(str(amount))

                time.sleep(0.5)
                driver.find_element(By.XPATH, '//*[@id="WRK_HGH"]').click()
                WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                if apcs1 != "x":
                    driver.find_element(By.XPATH, '//*[@id="pcs_tr"]/td/div[1]/label[5]/div[1]/img').click()
                    time.sleep(0.5)

                WebDriverWait(driver, 5).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                driver.find_element(By.XPATH, '//*[@id="WRK_HGH"]').click()
                total_price = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.ID, "PRICE_DIS2")))
                # driver.execute_script("$('#chk_notice_confirm').click();")  # 확인했습니다.
                # time.sleep(2)

                tprice = total_price.text
                time.sleep(0.5)
                # if str(amount) == '1':
                #     print(i, "번째")
                #     time.sleep(0.5)
                # else:
                driver.find_element(By.ID, 'direct_order_btn').click()
                # # time.sleep(3)
                # # al = Alert(driver)
                # # al.accept()
                # # time.sleep(0.5)
                imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
                print(i, "번째 TOTAL_PRICE : ", tprice, "pot_tmp_cod : ", imsiordernum)
                driver.execute_script('window.location.reload();')
                driver.execute_script('window.scrollTo(0, 100);')
                time.sleep(2)
        elif itemCode in itemList_offset:
            for i in range(totalList):
                paper = df_item.iloc[i, 1]
                wgtcod = df_item.iloc[i, 2]
                dosu = df_item.iloc[i, 3]
                size = df_item.iloc[i, 4]
                amount = df_item.iloc[i, 5]
                apcs1 = df_item.iloc[i, 6]
                apcs2 = df_item.iloc[i, 7]
                apcs3 = df_item.iloc[i, 8]
                apcs4 = df_item.iloc[i, 9]
                apcs5 = df_item.iloc[i, 10]
                time.sleep(2)
                # 용지 선택
                paper_select = driver.find_element(By.XPATH, '//*[@id="app"]/section/section[1]/div[3]/div/select[1]')
                pselect = Select(paper_select)
                for paperoption in pselect.options:
                    if paper == paperoption.text:
                        pselect.select_by_visible_text(paperoption.text)

                # 용지무게 선택
                time.sleep(1)
                wgtcod_select = driver.find_element(By.XPATH, '//*[@id="app"]/section/section[1]/div[3]/div/select[2]')
                wselect = Select(wgtcod_select)
                for wgtcodoption in wselect.options:
                    if wgtcod == wgtcodoption.text:
                        wselect.select_by_visible_text(wgtcodoption.text)

                # dosu : 인쇄도수 선택
                time.sleep(1)
                dosu_select = driver.find_element(By.XPATH, '//*[@id="app"]/section/section[1]/div[5]/div/select')
                dselect = Select(dosu_select)
                for dosuoption in dselect.options:
                    if dosu == dosuoption.text:
                        dselect.select_by_visible_text(dosuoption.text)

                # size : 사이즈 선택
                time.sleep(1)
                size_select = driver.find_element(By.XPATH, '//*[@id="app"]/section/section[1]/div[4]/div[1]/select')
                size_value = size.replace("*", "X")
                if itemCode == 'NCCDDFT':
                    size_value = size
                sselect = Select(size_select)
                for sizeoption in sselect.options:
                    if size_value == sizeoption.text:
                        sselect.select_by_visible_text(sizeoption.text)

                amoun_btn = driver.find_element(By.XPATH, '//*[@id="app"]/section/section[1]/div[6]/div[2]/button')
                amoun_btn.click()

                amount_inputbtn = driver.find_element(By.XPATH, '//*[@id="app"]/section/section[1]/div[6]/div[2]/div[3]/input')
                amount_inputbtn.click()
                amount_inputbtn.send_keys(Keys.DELETE)
                amount_inputbtn.send_keys(Keys.DELETE)
                amount_inputbtn.send_keys(Keys.DELETE)
                amount_inputbtn.send_keys(Keys.DELETE)
                amount_inputbtn.send_keys(Keys.BACK_SPACE)
                amount_inputbtn.send_keys(Keys.BACK_SPACE)
                amount_inputbtn.send_keys(Keys.BACK_SPACE)
                amount_inputbtn.send_keys(Keys.BACK_SPACE)
                time.sleep(0.5)
                amount_inputbtn.send_keys(str(amount))

                order_btn = driver.find_element(By.XPATH, '//*[@id="app"]/section/section[1]/div[7]/div/div/input')
                order_btn.click()
                time.sleep(0.5)
                driver.execute_script('window.scrollTo(0, 1000);')
                if apcs1 == 'x':
                    time.sleep(0.5)
                else:
                    pcs1_img = driver.find_element(By.XPATH, '//*[@id="app"]/section/section[1]/div[8]/div/ul/li[1]/div/img')
                    pcs1_img.click()
                    if apcs1 == '무광':
                        time.sleep(0.5)
                    else:
                        coting = driver.find_element(By.XPATH, '//*[@id="app"]/section/section[1]/div[8]/div/div[1]/div/ul[2]/li[2]/div/img')
                        coting.click()
                        #무광 유광

                time.sleep(1)
                driver.execute_script('window.scrollTo(0, 1000);')
                # driver.find_element(By.ID, 'direct_order_btn').click()
                order_btn = driver.find_element(By.XPATH, '//*[@id="app"]/section/section[1]/div[7]/div/div/input')
                order_btn.click()
                total_price = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "PRICE_DIS2")) )
                tprice = total_price.text
                time.sleep(1)
                imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
                if imsiordernum == '':
                    driver.find_element(By.ID, 'direct_order_btn').click()

                time.sleep(3)
                al = Alert(driver)
                al.accept()
                time.sleep(1)
                # imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
                print(i, "번째 TOTAL_PRICE : ", tprice, "pot_tmp_cod : ", imsiordernum)
                # print(i, "번째 TOTAL_PRICE : ", tprice)
                driver.execute_script('window.scrollTo(0, 100);')
                time.sleep(2)
        elif itemCode in itemList_memo:
            for i in range(totalList):
                time.sleep(2)
                dosu = df_item.iloc[i, 3]
                size = df_item.iloc[i, 4]
                amount = df_item.iloc[i, 5]
                apcs1 = df_item.iloc[i, 6]
                apcs2 = df_item.iloc[i, 7]
                apcs3 = df_item.iloc[i, 8]

                driver.execute_script('window.scrollTo(0, 500);')
                time.sleep(0.5)

                driver.execute_script('$("#'+size+'").click();')
                time.sleep(0.5)
                driver.execute_script('$("#'+size+'").click();')
                time.sleep(1)

                driver.execute_script('$("#'+dosu+'").click();')
                time.sleep(0.5)
                driver.execute_script('$("#'+dosu+'").click();')
                time.sleep(1)

                driver.execute_script("productOrder.check_PRN_CNT();")
                number1_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "number1")) )
                if number1_text.get_attribute('value') != str(amount):
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'number1').click()
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'number1').send_keys(str(amount))

                time.sleep(0.5)
                driver.execute_script('$("#'+apcs1+'").click();')
                time.sleep(0.5)
                driver.execute_script('$("#'+apcs1+'").click();')
                time.sleep(0.5)

                driver.execute_script('$("#'+apcs2+'").click();')
                time.sleep(0.5)
                driver.execute_script('$("#'+apcs2+'").click();')
                time.sleep(0.5)

                # time.sleep(3)
                # next_btn = driver.find_element(By.XPATH, '//*[@id="price_info"]/div[3]/div[1]/input')
                # next_btn.click()

                # time.sleep(1)
                # time.sleep(0.5)
                # driver.execute_script('$("#btn-confirm").click();')
                # time.sleep(2)

                # driver.execute_script('$("#chk_notice_confirm").click();')
                # time.sleep(1)

                total_price = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "PRICE_DIS2")) )
                tprice = total_price.text
                # time.sleep(1)
                # driver.find_element(By.ID, 'direct_order_btn').click()
                # time.sleep(5)
                # driver.find_element(By.ID, 'direct_order_btn').click()
                # time.sleep(2)
                # al = Alert(driver)
                # al.accept()
                # time.sleep(0.5)
                # imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
                print(i, "번째 TOTAL_PRICE : ", tprice) #, "pot_tmp_cod : ", imsiordernum
                # print(i, "번째 TOTAL_PRICE : ", tprice)
                driver.execute_script('window.location.reload();')
                driver.execute_script('window.scrollTo(0, 100);')
                time.sleep(2)
        elif itemCode in itemList_kring:
            wait = WebDriverWait(driver, 10)
            for i in range(totalList):
                time.sleep(2)
                colum_1 = df_item.iloc[i, 1] #weight
                colum_2 = df_item.iloc[i, 2] #size
                colum_3 = df_item.iloc[i, 3] #color
                colum_4 = df_item.iloc[i, 4] #sheet
                colum_5 = df_item.iloc[i, 5] #sheetColor
                colum_6 = df_item.iloc[i, 6] #gori
                colum_7 = df_item.iloc[i, 7] #dosu
                colum_8 = df_item.iloc[i, 8] #cnt

                # driver.execute_script("arguments[0].style.display = 'block';", driver.find_element(By.ID, "paper_tr"))
                driver.execute_script('window.scrollTo(0, 100);')
                time.sleep(0.5)

                weight_select = wait.until( EC.visibility_of_element_located((By.ID, "strap_size_hgh")) )
                weight = Select(weight_select)
                weight.select_by_visible_text(colum_1)

                if colum_1 == '두께: 25(mm)':
                    size_select = wait.until(EC.visibility_of_element_located((By.ID, "strap_size_25")))
                    size = Select(size_select)
                    size.select_by_visible_text(colum_2)
                else:
                    size_select = wait.until(EC.visibility_of_element_located((By.ID, "strap_size_30")))
                    size = Select(size_select)
                    size.select_by_visible_text(colum_2)

                color_select = wait.until( EC.visibility_of_element_located((By.ID, "strap_cmc_clr")) )
                color = Select(color_select)
                color.select_by_visible_text(colum_3)

                sheet_select = wait.until( EC.visibility_of_element_located((By.ID, "strap_paper")) )
                sheet = Select(sheet_select)
                sheet.select_by_visible_text(colum_4)

                time.sleep(1)
                driver.execute_script('productOrder.strap_paper2("'+colum_5+'");')

                sheet_select = wait.until( EC.visibility_of_element_located((By.ID, "strap_ring")) )
                sheet = Select(sheet_select)
                sheet.select_by_visible_text(colum_6)

                docu_text = wait.until( EC.visibility_of_element_located((By.ID, "soduSelectBoxItText")))
                if docu_text.text != colum_7:
                    driver.find_element(By.ID, 'soduSelectBoxIt').click()
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    driver.find_element(By.LINK_TEXT, colum_7).click()

                driver.execute_script('window.scrollTo(0, 400);')
                # total_price = wait.until( EC.visibility_of_element_located((By.ID, "PRICE_DIS2")) )
                # tprice = total_price.text
                time.sleep(1)
                try:
                     driver.find_element(By.ID, 'direct_order_btn').click()
                     try:
                         # alert가 뜰 때까지 기다리기 (최대 3초)
                         alert = WebDriverWait(driver, 3).until(EC.alert_is_present())
                         alert.accept()
                         wait.until(EC.element_to_be_clickable((By.ID, 'direct_order_btn'))).click()
                     except TimeoutException:
                         time.sleep(1)

                except Exception as e:
                    print("예외 발생:", e)
                # time.sleep(5)
                # driver.find_element(By.ID, 'direct_order_btn').click()
                time.sleep(2)
                print(i, "번째 : ", colum_1, "|", colum_2, "|", colum_3, "|", colum_4, "|", colum_5, "|", colum_6, "|",
                      colum_7)  # , "pot_tmp_cod : ", imsiordernum
                # al = Alert(driver)
                # al.accept()
                # time.sleep(0.5)
                # imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
                # print(i, "번째 : ", colum_1, "|", colum_2, "|", colum_3, "|", colum_4, "|", colum_5, "|", colum_6, "|", colum_7) #, "pot_tmp_cod : ", imsiordernum
                # print(i, "번째 TOTAL_PRICE : ", tprice)
                driver.execute_script('window.location.reload();')
                driver.execute_script('window.scrollTo(0, 100);')
                time.sleep(2)
        elif itemCode in itemList_pendunte:
            wait = WebDriverWait(driver, 10)
            for i in range(totalList):
                time.sleep(2)
                colum_1 = df_item.iloc[i, 1] #규격
                colum_2 = df_item.iloc[i, 2] #사이즈
                colum_3 = df_item.iloc[i, 3] #인쇄도수
                colum_4 = df_item.iloc[i, 4] #목걸이
                colum_5 = df_item.iloc[i, 5] #목걸이두께
                colum_6 = df_item.iloc[i, 6] #개별포장여부

                # driver.execute_script("arguments[0].style.display = 'block';", driver.find_element(By.ID, "paper_tr"))
                driver.execute_script('window.scrollTo(0, 100);')
                time.sleep(0.5)

                driver.execute_script("productOrder.check_GSPASGC('shape_type', '"+colum_1+"');")

                time.sleep(1)
                size_select = wait.until( EC.visibility_of_element_located((By.ID, "pendant_size")) )
                size = Select(size_select)
                size.select_by_visible_text(colum_2)

                docu_text = wait.until( EC.visibility_of_element_located((By.ID, "soduSelectBoxItText")))
                if docu_text.text != colum_3:
                    driver.find_element(By.ID, 'soduSelectBoxIt').click()
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    driver.find_element(By.LINK_TEXT, colum_3).click()

                time.sleep(0.5)
                if colum_4 == 'x':
                    time.sleep(0.5)
                else:
                    driver.execute_script('$("#'+colum_4+'_chain_CHK").click();')
                    time.sleep(0.5)
                    driver.execute_script('$("#'+colum_4+'_chain_CHK").click();')

                time.sleep(0.5)
                if colum_5 == 'x':
                    time.sleep(0.5)
                else:
                    driver.execute_script('$("#'+colum_5+'_thickness_CHK").click();')
                    time.sleep(0.5)
                    driver.execute_script('$("#'+colum_5+'_thickness_CHK").click();')
                    # driver.execute_script(" productOrder.check_GSPASGC('thickness', '"+colum_5+"'); ")

                time.sleep(0.5)
                if colum_6 == 'x':
                    time.sleep(0.5)
                else:
                    driver.execute_script("  productOrder.opt_use_yn('"+colum_6+"', '');  ")


                driver.execute_script('window.scrollTo(0, 400);')
                # total_price = wait.until( EC.visibility_of_element_located((By.ID, "PRICE_DIS2")) )
                # tprice = total_price.text
                time.sleep(1)
                driver.execute_script("$('#chk_notice_confirm').click();")  # 확인했습니다.
                time.sleep(1)
                try:
                     driver.find_element(By.ID, 'direct_order_btn').click()
                     try:
                         # alert가 뜰 때까지 기다리기 (최대 3초)
                         alert = WebDriverWait(driver, 3).until(EC.alert_is_present())
                         alert.accept()
                         wait.until(EC.element_to_be_clickable((By.ID, 'direct_order_btn'))).click()
                     except TimeoutException:
                         time.sleep(1)

                except Exception as e:
                    print("예외 발생:", e)
                # time.sleep(5)
                # driver.find_element(By.ID, 'direct_order_btn').click()
                time.sleep(1)
                print(i, "번째 : ", colum_1, "|", colum_2, "|", colum_3, "|", colum_4, "|", colum_5, "|", colum_6)  # , "pot_tmp_cod : ", imsiordernum
                # al = Alert(driver)
                # al.accept()
                # time.sleep(1)
                # imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
                # print(i, "번째 : ", colum_1, "|", colum_2, "|", colum_3, "|", colum_4, "|", colum_5, "|", colum_6, "|", colum_7) #, "pot_tmp_cod : ", imsiordernum
                # print(i, "번째 TOTAL_PRICE : ", tprice)
                driver.execute_script('window.location.reload();')
                driver.execute_script('window.scrollTo(0, 100);')
                time.sleep(1)
        elif itemCode in itemList_PRCAXPO:
            wait = WebDriverWait(driver, 5)
            for i in range(totalList):
                time.sleep(2)
                colum_1 = df_item.iloc[i, 1] #용지
                colum_2 = df_item.iloc[i, 2] #용지무게
                colum_3 = df_item.iloc[i, 3] #가로세로
                colum_4 = df_item.iloc[i, 4] #인쇄도수
                colum_5 = df_item.iloc[i, 5] #코팅
                colum_6 = df_item.iloc[i, 6] #코팅상세
                colum_7 = df_item.iloc[i, 7] #화이트인쇄
                colum_8 = df_item.iloc[i, 8] #수량

                # driver.execute_script("arguments[0].style.display = 'block';", driver.find_element(By.ID, "paper_tr"))
                driver.execute_script('window.scrollTo(0, 200);')
                time.sleep(0.5)

                if colum_3 == '가로':
                    driver.execute_script("productOrder.paper_wh('W');")
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                elif colum_3 == '세로':
                    driver.execute_script("productOrder.paper_wh('H');")
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                else:
                    time.sleep(0.5)

                wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                paper_text = wait.until(EC.visibility_of_element_located((By.ID, "paperSelectBoxItText")))
                if paper_text.text != colum_1:
                    # paperSelectBoxItContainer
                    driver.find_element(By.ID, 'paperSelectBoxIt').click()
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    driver.find_element(By.LINK_TEXT, colum_1).click()

                wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                wgt_text = wait.until(EC.visibility_of_element_located((By.ID, "paper_sub_selectSelectBoxItText")))
                if wgt_text.text != str(colum_2):
                    # paper_sub_selectSelectBoxItContainer
                    driver.find_element(By.ID, 'paper_sub_selectSelectBoxIt').click()
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    driver.find_element(By.LINK_TEXT, str(colum_2)).click()

                wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                docu_text = wait.until( EC.visibility_of_element_located((By.ID, "soduSelectBoxItText")))
                if docu_text.text != colum_4:
                    driver.find_element(By.ID, 'soduSelectBoxIt').click()
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    driver.find_element(By.LINK_TEXT, colum_4).click()

                wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                driver.execute_script("productOrder.check_PRN_CNT();")
                number1_text = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.ID, "number1")))
                if number1_text.get_attribute('value') != str(colum_8):
                    time.sleep(1)
                    driver.find_element(By.ID, 'number1').click()
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    time.sleep(1)
                    driver.find_element(By.ID, 'number1').send_keys(str(colum_8))

                driver.find_element(By.ID, 'WRK_HGH').click()
                # time.sleep(10)
                if colum_5 == 'x':
                    time.sleep(0.5)
                else:
                    driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    time.sleep(0.5)
                    if colum_5 == '양면':
                        driver.execute_script('$("#COT_DFT_RADIO_D").click();')
                        wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                        time.sleep(0.5)
                    else:
                        time.sleep(0.5)

                if colum_6 == '유광':
                    driver.execute_script('$("#COT_DFT_SUB_RADIO_GL").click();')
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                else:
                    time.sleep(0.5)

                driver.find_element(By.ID, 'WRK_HGH').click()
                if colum_7 == 'x':
                    time.sleep(0.5)
                else:
                    driver.execute_script("productOrder.opt_use_yn('PRT_WHT', '     ');")
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    time.sleep(0.5)

                driver.find_element(By.ID, 'WRK_HGH').click()
                # driver.execute_script("productOrder.order_validate('pot_create');")  # 확인했습니다.
                # time.sleep(1)
                try:
                    driver.execute_script("productOrder.order_validate('pot_create');")
                    time.sleep(1)
                    # driver.find_element(By.ID, 'direct_order_btn').click()
                    # try:
                    #     # alert가 뜰 때까지 기다리기 (최대 3초)
                    #     alert = wait.until(EC.alert_is_present())
                    #     alert.accept()
                    #     wait.until(EC.element_to_be_clickable((By.ID, 'direct_order_btn'))).click()
                    # except TimeoutException:
                    #     wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))

                except Exception as e:
                    print("예외 발생:", e)
                # time.sleep(5)
                # driver.find_element(By.ID, 'direct_order_btn').click()
                # time.sleep(2)
                # imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
                print(i, "번째 : ", colum_1, "|", colum_2, "|", colum_3, "|", colum_4, "|", colum_5, "|", colum_6, "|", colum_7, "|", colum_8)  # , "pot_tmp_cod : ", imsiordernum
                # print(i, "번째 : ", colum_1, "|", colum_2, "|", colum_3, "|", colum_4, "|", colum_5, "|", colum_6, "|", colum_7) #, "pot_tmp_cod : ", imsiordernum
                # print(i, "번째 TOTAL_PRICE : ", tprice)
                driver.execute_script('window.location.reload();')
                time.sleep(2)
        elif itemCode in itemList_LFXXXXX:
            wait = WebDriverWait(driver, 5)
            for i in range(totalList):
                time.sleep(2)
                colum_1 = df_item.iloc[i, 1] #용지
                colum_2 = df_item.iloc[i, 2] #용지무게
                colum_3 = df_item.iloc[i, 3] #인쇄도수
                colum_4 = df_item.iloc[i, 4] #사이즈
                colum_5 = df_item.iloc[i, 5] #수량
                colum_6 = df_item.iloc[i, 6] #코팅
                colum_9 = df_item.iloc[i, 9] #재단
                colum_10 = df_item.iloc[i, 10] #화이트인쇄

                # driver.execute_script("arguments[0].style.display = 'block';", driver.find_element(By.ID, "paper_tr"))
                driver.execute_script('window.scrollTo(0, 200);')
                time.sleep(0.5)
                if itemCode == 'STCUXXX':
                    sizesplit = colum_4.split("*")
                    wid_size = sizesplit[0]
                    hei_size = sizesplit[1]
                    CUT_WDT_text = wait.until( EC.visibility_of_element_located((By.ID, "CUT_WDT")) )
                    if CUT_WDT_text.get_attribute('value') != wid_size:
                        # 상품 페이지 사이즈 직접입력 사이즈 입력
                        driver.find_element(By.ID, 'CUT_WDT').click()
                        time.sleep(0.5)
                        driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                        time.sleep(0.5)
                        driver.find_element(By.ID, 'CUT_WDT').send_keys(wid_size)

                    CUT_HGH_text = wait.until( EC.visibility_of_element_located((By.ID, "CUT_HGH")) )
                    if CUT_HGH_text.get_attribute('value') != hei_size:
                        driver.find_element(By.ID, 'CUT_HGH').click()
                        time.sleep(0.5)
                        driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                        time.sleep(0.5)
                        driver.find_element(By.ID, 'CUT_HGH').send_keys(hei_size)

                else:
                    size_text = wait.until(EC.visibility_of_element_located((By.ID, "sizeSelectBoxItText")))
                    if size_text.text != colum_4:
                        # 사이즈 직접입력 선택 sizeSelectBoxItContainer
                        driver.find_element(By.ID, 'sizeSelectBoxIt').click()
                        wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                        driver.find_element(By.LINK_TEXT, colum_4).click()

                wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                paper_text = wait.until(EC.visibility_of_element_located((By.ID, "paperSelectBoxItText")))
                if paper_text.text != colum_1:
                    # paperSelectBoxItContainer
                    driver.find_element(By.ID, 'paperSelectBoxIt').click()
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    driver.find_element(By.LINK_TEXT, colum_1).click()

                wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                wgt_text = wait.until(EC.visibility_of_element_located((By.ID, "paper_sub_selectSelectBoxItText")))
                if wgt_text.text != str(colum_2):
                    # paper_sub_selectSelectBoxItContainer
                    driver.find_element(By.ID, 'paper_sub_selectSelectBoxIt').click()
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    driver.find_element(By.LINK_TEXT, str(colum_2)).click()

                wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                docu_text = wait.until( EC.visibility_of_element_located((By.ID, "soduSelectBoxItText")))
                if docu_text.text != colum_3:
                    driver.find_element(By.ID, 'soduSelectBoxIt').click()
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    driver.find_element(By.LINK_TEXT, colum_3).click()

                wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                driver.execute_script("productOrder.check_PRN_CNT();")
                number1_text = wait.until(EC.visibility_of_element_located((By.ID, "number1")))
                if number1_text.get_attribute('value') != str(colum_5):
                    time.sleep(1)
                    driver.find_element(By.ID, 'number1').click()
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    time.sleep(1)
                    driver.find_element(By.ID, 'number1').send_keys(str(colum_5))

                driver.find_element(By.ID, 'WRK_HGH').click()
                print(i, "번째 : ", colum_1, "|", colum_2, "|", colum_3, "|", colum_4, "|", colum_5, "|", colum_6, "|", colum_9, "|", colum_10)  # , "pot_tmp_cod : ", imsiordernum, "|", colum_7, "|", colum_8
                if colum_6 == 'x':
                    time.sleep(0.5)
                elif colum_6 == '무광단면':
                    driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                elif colum_6 == '무광':
                    driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                elif colum_6 == '유광단면':
                    driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    time.sleep(0.5)
                    driver.execute_script("productOrder.opt_select('COT_DFT','GL');")
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                elif colum_6 == '유광':
                    driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    time.sleep(0.5)
                    driver.execute_script("productOrder.opt_select('COT_DFT','GL');")
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                elif colum_6 == '무광양면':
                    driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    time.sleep(0.5)
                    driver.execute_script('$("#COT_DFT_RADIO_D").click();')
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    driver.execute_script('$("#COT_DFT_RADIO_D").click();')
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                elif colum_6 == '유광양면':
                    driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    time.sleep(0.5)
                    driver.execute_script('$("#COT_DFT_RADIO_D").click();')
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    driver.execute_script('$("#COT_DFT_RADIO_D").click();')
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    driver.execute_script("productOrder.opt_select('COT_DFT','GL');")
                    wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                else:
                    time.sleep(10)

                if colum_9 == "묶음재단":
                    driver.execute_script("productOrder.opt_checked('CUT_DFT', 'DFXXX');")  # 상품 페이지 묶음재단 선택
                elif colum_9 == "개별재단":
                    driver.execute_script("productOrder.opt_checked('CUT_DFT', 'DFITM');")  # 상품 페이지 개별재단 선택
                else:
                    time.sleep(0.2)

                driver.find_element(By.ID, 'WRK_HGH').click()
                if colum_10 == 'x':
                    time.sleep(0.5)
                else:
                    if colum_1 == '금광 PET':
                        driver.execute_script("productOrder.opt_use_yn('PRT_WHT2', '     ');")
                        wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                        time.sleep(0.5)
                    elif colum_1 == '은광 PET':
                        driver.execute_script("productOrder.opt_use_yn('PRT_WHT2', '     ');")
                        wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                        time.sleep(0.5)
                    elif colum_1 == '고투명 PET 리무버블':
                        driver.execute_script("productOrder.opt_use_yn('PRT_WHT2', '     ');")
                        wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                        time.sleep(0.5)
                    else:
                        driver.execute_script("productOrder.opt_use_yn('PRT_WHT', '     ');")
                        wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                        time.sleep(0.5)

                time.sleep(0.5)
                # if colum_5 == 'x':
                #     time.sleep(0.5)
                # else:
                #     driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                #     wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                #     time.sleep(0.5)
                #     if colum_5 == '양면':
                #         driver.execute_script('$("#COT_DFT_RADIO_D").click();')
                #         wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                #         time.sleep(0.5)
                #     else:
                #         time.sleep(0.5)
                #
                # if colum_6 == '유광':
                #     driver.execute_script('$("#COT_DFT_SUB_RADIO_GL").click();')
                #     wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                # else:
                #     time.sleep(0.5)


                # driver.find_element(By.ID, 'WRK_HGH').click()
                # driver.execute_script("productOrder.order_validate('pot_create');")  # 확인했습니다.
                # time.sleep(1)
                try:
                    driver.execute_script("productOrder.order_validate('pot_create');")
                    time.sleep(1)
                    # driver.find_element(By.ID, 'direct_order_btn').click()
                    # try:
                    #     # alert가 뜰 때까지 기다리기 (최대 3초)
                    #     alert = wait.until(EC.alert_is_present())
                    #     alert.accept()
                    #     wait.until(EC.element_to_be_clickable((By.ID, 'direct_order_btn'))).click()
                    # except TimeoutException:
                    #     wait.until(EC.invisibility_of_element_located((By.ID, 'overlay')))

                except Exception as e:
                    print("예외 발생:", e)
                # time.sleep(5)
                # driver.find_element(By.ID, 'direct_order_btn').click()
                # time.sleep(2)
                # imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
                # print(i, "번째 : ", colum_1, "|", colum_2, "|", colum_3, "|", colum_4, "|", colum_5, "|", colum_6, "|", colum_7) #, "pot_tmp_cod : ", imsiordernum
                print('생성완료.')
                driver.execute_script('window.location.reload();')
                time.sleep(2)

        else:
            print('error')

        # nowtime=str(now.year)+'_'+str(now.month)+'_'+str(now.day)+'_'+str(now.hour)+'_'+str(now.minute)
        # new_filename = REDATA_FOLDER+'/'+userid+'_'+itemCode+'_' + nowtime + '.txt'
        # with open(new_filename, 'w', encoding='ansi') as file:
        #     file.write(newFileData)
        # # df_item.to_excel(new_filename, sheet_name=itemCode, index=False)
        # time.sleep(5)
        # driver.quit()
    except WebDriverException as e:
        print(f"WebDriver 오류 발생: {e}")
        # login_check_proc(userid, userpw, itemUrl, driver)
        # driver.quit()

    finally:
        print(f"finally 오류 발생:")
        # driver.quit()

def uploadfile_ordernum_creating_pan(df_item_config, df_item):
    totalList = len(df_item)
    userid = df_item_config.iloc[0,2]
    userpw = df_item_config.iloc[0,3]
    if userid == "x":
        userid = 'red_openmarket' #red_openmarket, #redprinting
    if userpw == "x":
        userpw = 'red4874#' #red4874# , #redprinting#1234

    options = ChromeOptions()
    options.add_argument('--blink-settings=imagesEnabled=false')
    driver = webdriver.Chrome(options=options)
    driver.implicitly_wait(3)
    for i in range(totalList):
        time.sleep(2)
        return_file = fileCreaft_orderConNumCreate(userid, userpw, df_item, driver, i)
        time.sleep(2)

        # try:
        #     # if (i + 1) % 5 == 0:
        #     #     nowtime = str(now.year) + '' + str(now.month) + '' + str(now.day) + '_' + str(now.hour) + '' + str(
        #     #         now.minute)
        #     #     new_filename = REDATA_FOLDER + '/' + userid + '_' + df_item.iloc[0, 0] + "_" + str(i + 1) + '_' + nowtime + '.txt'
        #     #     print(new_filename);
        #         # with open(new_filename, 'w', encoding='ansi') as file:
        #         #     file.write(return_file)
        #
        # except Exception as e:
        #     print(f"Error: {e}")
        #     time.sleep(3)
        #     return_file = fileCreaft_orderConNumCreate(userid, userpw, df_item, driver, i)
        #     time.sleep(2)
        #     # if (i + 1) % 5 == 0:
        #     #     nowtime = str(now.year) + '' + str(now.month) + '' + str(now.day) + '_' + str(now.hour) + '' + str(
        #     #         now.minute)
        #     #     new_filename = REDATA_FOLDER + '/' + userid + '_' + df_item.iloc[0, 0] + "_" + str(i + 1) + '_' + nowtime + '.txt'
        #     #     print(new_filename);
        #         # with open(new_filename, 'w', encoding='ansi') as file:
        #         #     file.write(return_file)
        #     # driver.quit()
        # finally:
        #     print("finally")
        #     # driver.quit()

        # time.sleep(5)
        # driver.quit()

def fileCreaft_orderConNumCreate(userid, userpw, df_item, driver, i):
    # 1~7
    newFileData = "아이템코드\t링크\t용지\t용지무게\t수량\t코팅후가공\t주문관리코드\t금액\n"
    itemCode = df_item.iloc[i, 0]
    itemUrl = df_item.iloc[i, 1]
    paper = df_item.iloc[i, 2]
    wgtcod = df_item.iloc[i, 3]
    amount = df_item.iloc[i, 4]
    apcs1 = df_item.iloc[i, 5]
    if i == 0:
        login_check_proc(userid, userpw, itemUrl, driver)
    else:
        time.sleep(3)
        driver.get(itemUrl)

    # newFileData = newFileData + itemCode + "\t" + str(itemUrl) + "\t" + str(paper) + "\t" + str(wgtcod) + "\t" + str(
    #     amount) + "\t" + str(apcs1) + "\t"
    time.sleep(3)
    if itemCode == 'BNSTDGN':
        time.sleep(0.5)
    else:
        paper_text = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.ID, "paperSelectBoxItText")))
        if paper_text.text != paper:
            # 상품 페이지 용지선택
            WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
            # paperSelectBoxItContainer
            driver.find_element(By.ID, 'paperSelectBoxIt').click()
            WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
            driver.find_element(By.LINK_TEXT, paper).click()

        WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
        if wgtcod != "":
            # 상품 페이지 G수 선택
            wgt_text = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.ID, "paper_sub_selectSelectBoxItText")))
            if wgt_text.text != str(wgtcod):
                # paper_sub_selectSelectBoxItContainer
                driver.find_element(By.ID, 'paper_sub_selectSelectBoxIt').click()
                WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                driver.find_element(By.LINK_TEXT, str(wgtcod)).click()

        WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
        driver.execute_script("productOrder.check_PRN_CNT();")
        number1_text = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.ID, "number1")))
        if number1_text.get_attribute('value') != str(amount):
            time.sleep(0.5)
            driver.find_element(By.ID, 'number1').click()
            driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
            driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
            driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
            driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
            driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
            driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
            driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
            driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
            time.sleep(0.5)
            driver.find_element(By.ID, 'number1').send_keys(str(amount))

    WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
    apcs_nowstatus = driver.find_element(By.ID, 'priceCalcResult').get_attribute('value')
    if apcs1 == "무광":
        if 'COT_DFT' in apcs_nowstatus:
            time.sleep(0.2)
        else:
            # 상품 페이지 코팅 선택
            driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
        WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
        driver.execute_script("productOrder.opt_select('COT_DFT','MA');")  # 상품 페이지 무광코팅 선택
    elif apcs1 == "유광":
        if 'COT_DFT' in apcs_nowstatus:
            time.sleep(0.2)
        else:
            # 상품 페이지 코팅 선택
            driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
        WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
        driver.execute_script("productOrder.opt_select('COT_DFT','GL');")  # 상품 페이지 유광코팅 선택
    else:
        if 'COT_DFT' in apcs_nowstatus:
            driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")  # 코팅 후가공 다시 선택 시 선택해제됨.
        else:
            time.sleep(0.2)

    time.sleep(0.5)
    driver.execute_script('window.scrollTo(0, 300)')
    time.sleep(0.5)

    WebDriverWait(driver, 5).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
    time.sleep(0.5)
    total_price = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.ID, "TOTAL_PRICE")))
    tprice = total_price.text
    time.sleep(0.5)
    driver.execute_script("productOrder.order_validate('pot_create');")
    # driver.find_element(By.ID, 'direct_order_btn').click()
    # time.sleep(3)
    # driver.execute_script('$("#overlay").css("display", "none");')
    time.sleep(4)
    al = Alert(driver)
    al.accept()
    time.sleep(0.5)
    # imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
    # imsiordernum = ""
    # print(i, "번째 TOTAL_PRICE : ", tprice, "pot_tmp_cod : ", imsiordernum)
    print(i, "번째 TOTAL_PRICE : ", tprice)
    # newFileData = newFileData + "\t" + imsiordernum + "\t" + tprice + "\n"
    # df_item.loc[i, 'OrderCode'] = imsiordernum  # 주문관리코드생성후추가

    return newFileData;


if __name__ == '__main__':
    app3.run(debug=True, host='0.0.0.0', port=5002)
