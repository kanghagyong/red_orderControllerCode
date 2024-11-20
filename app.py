from flask import Flask, send_from_directory, render_template_string
import os
import time
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
# import requests
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By
# 웹드라이버 생성
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.common.exceptions import WebDriverException
now = datetime.now()

pd.set_option('display.width', 320)
pd.set_option('display.max_columns', 20)
import tempfile

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
LIST_FOLDER = 'list'
DATA_FOLDER = 'data'
REDATA_FOLDER = 'reData'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route('/')
def index():
    xlsx_files = [f for f in os.listdir(DATA_FOLDER) if f.endswith('.xlsx')]
    file_links = ''.join(f'<li><a href="/download/{f}">{f}</a></li>' for f in xlsx_files)
    first_fileupload =  render_template_string('''
        <div style="width:50%;float:left;">
            <h1>옵션만들기 파일업로드</h1>
            <form action="/upload" method="post" enctype="multipart/form-data">
                <input type="file" name="file"><br><br>
                <input type="submit" value="파일생성">
            </form>
            <h1>다운로드(생성된옵션리스트) 목록</h1>
            <ul>
                {{ files|safe }}
            </ul>
        </div>
    ''', files=file_links)

    xlsx_files1 = [f for f in os.listdir(REDATA_FOLDER) if f.endswith('.xlsx')]
    file_links1 = ''.join(f'<li><a href="/download2/{f}">{f}</a></li>' for f in xlsx_files1)
    first_fileupload1 =  render_template_string('''
        <div style="width:50%;float:left;">
            <h1>주문관리코드생성 파일업로드</h1>
            <form action="/upload2" method="post" enctype="multipart/form-data">
                <input type="file" name="file"><br><br>
                <input type="submit" value="파일생성">
            </form>
            <h1>다운로드(생성된주문관리코드) 목록</h1>
            <ul>
                {{ files1|safe }}
            </ul>
        </div>
    ''', files1=file_links1)

    return first_fileupload+first_fileupload1

@app.route('/upload', methods=['POST'])
def upload_file():
    from flask import request

    if 'file' not in request.files:
        return '파일이 없습니다.'
    file = request.files['file']
    if file.filename == '':
        return '파일 이름이 없습니다.'

    file.save(os.path.join(UPLOAD_FOLDER, file.filename))

    start_filename = UPLOAD_FOLDER+'/'+file.filename
    df_item_config = pd.read_excel(start_filename, sheet_name=0, engine='openpyxl', skiprows=1)
    uploadfile_option_check(df_item_config)
    time.sleep(3)
    return '파일이 업로드되었습니다!<br><a href="/">목록</a>'

@app.route('/upload2', methods=['POST'])
def upload2_file():
    from flask import request

    if 'file' not in request.files:
        return '파일이 없습니다.'
    file = request.files['file']
    if file.filename == '':
        return '파일 이름이 없습니다.'

    file.save(os.path.join(UPLOAD_FOLDER, file.filename))

    start_filename = UPLOAD_FOLDER+'/'+file.filename
    df_item_config = pd.read_excel(start_filename, sheet_name=0, engine='openpyxl', nrows=1)
    df_item = pd.read_excel(start_filename, sheet_name=0, engine='openpyxl', skiprows=7)

    uploadfile_ordernum_creating(df_item_config, df_item)
    time.sleep(3)

    return '파일이 업로드되었습니다!<br><a href="/">목록</a>'

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

@app.route('/download/<filename>')
def download(filename):
    # temp_dir = tempfile.gettempdir()
    return send_from_directory(DATA_FOLDER, filename, as_attachment=True)

@app.route('/download2/<filename>')
def download2(filename):
    # temp_dir = tempfile.gettempdir()
    return send_from_directory(REDATA_FOLDER, filename, as_attachment=True)

def uploadfile_option_check(df_item_config):
    itemList_stationery = ['GSSBMTL', 'GSSBACM', 'GSSBSTP']
    totalList = len(df_item_config)
    nowtime = str(now.year) + '' + str(now.month) + '' + str(now.day) + '_' + str(now.hour) + '' + str(
        now.minute) + '' + str(now.second) + '' + str(now.microsecond)

    for item in range(totalList):
        # 상품코드 가져오기
        itemCode = df_item_config.iloc[item, 0]
        if itemCode in itemList_stationery:
            print('문방구 상품은 옵션 리스트를 따로 만들수 없습니다.')
            continue
        else:
            sampleOption_filename = LIST_FOLDER+'/option_sample.xlsx'
            df_item = pd.read_excel(sampleOption_filename, sheet_name=0, engine='openpyxl', skiprows=1)
            if "," in df_item_config.iloc[item, 1]:
                papers_wgtlist = df_item_config.iloc[item, 1].split(",")
            else:
                papers_wgtlist = []
                papers_wgtlist.append(df_item_config.iloc[item, 1])
            if "," in df_item_config.iloc[item, 2]:
                dosulist = df_item_config.iloc[item, 2].split(",")
            else:
                dosulist = []
                dosulist.append(df_item_config.iloc[item, 2])
            if "," in df_item_config.iloc[item, 3]:
                sizelist = df_item_config.iloc[item, 3].split(",")
            else:
                sizelist = []
                sizelist.append(df_item_config.iloc[item, 3])
            if "," in str(df_item_config.iloc[item, 4]):
                amountlist = df_item_config.iloc[item, 4].split(",")
            else:
                amountlist = []
                amountlist.append(str(df_item_config.iloc[item, 4]))
            if df_item_config.iloc[item, 5] == 'x':
                afterpcs01_list = ['x']
            else:
                afterpcs01_list = df_item_config.iloc[item, 5].split(",")

            if df_item_config.iloc[item, 6] == 'x':
                afterpcs02_list = ['x']
            else:
                afterpcs02_list = df_item_config.iloc[item, 6].split(",")

            if df_item_config.iloc[item, 7] == 'x':
                afterpcs03_list = ['x']
            else:
                afterpcs03_list = df_item_config.iloc[item, 7].split(",")

            if df_item_config.iloc[item, 8] == 'x':
                afterpcs04_list = ['x']
            else:
                afterpcs04_list = df_item_config.iloc[item, 8].split(",")

            if df_item_config.iloc[item, 9] == 'x':
                afterpcs05_list = ['x']
            else:
                afterpcs05_list = df_item_config.iloc[item, 9].split(",")

            # print(papers_wgtlist)
            # print(dosulist)
            # print(sizelist)
            # print(amountlist)
            # print(afterpcs01_list)
            # print(afterpcs02_list)
            # print(afterpcs03_list)
            # print(afterpcs04_list)
            # print(afterpcs05_list)

            papers = []  # 용지리스트
            wgt = []  # 용지무게 리스트
            for pw in papers_wgtlist:
                pwdata = pw.split("_")
                if pwdata[0] in papers:
                    time.sleep(0.2)
                else:
                    papers.append(pwdata)

            i = 0
            for p in papers:
                for d in dosulist:
                    for s in sizelist:
                        for a in amountlist:
                            for ap1 in afterpcs01_list:
                                # if ap1 == '유광':
                                for ap2 in afterpcs02_list:
                                    for ap3 in afterpcs03_list:
                                        for ap4 in afterpcs04_list:
                                            for ap5 in afterpcs05_list:
                                                df_item.loc[i, 'ItemCode'] = itemCode
                                                df_item.loc[i, 'Papers'] = p[0]
                                                df_item.loc[i, 'WgtCod'] = p[1]
                                                df_item.loc[i, 'Dosu'] = d
                                                df_item.loc[i, 'Sizes'] = s
                                                df_item.loc[i, 'Amount'] = a
                                                df_item.loc[i, 'AfterPcs01'] = ap1
                                                df_item.loc[i, 'AfterPcs02'] = ap2
                                                df_item.loc[i, 'AfterPcs03'] = ap3
                                                df_item.loc[i, 'AfterPcs04'] = ap4
                                                df_item.loc[i, 'AfterPcs05'] = ap5
                                                df_item.loc[i, 'OrderCode'] = ""
                                                df_item.loc[i, 'Price'] = ""
                                                i = i + 1

            print(i)
            new_filename = DATA_FOLDER+'/' + itemCode + '_' + nowtime + '.xlsx'
            df_item.to_excel(new_filename, sheet_name=itemCode, index=False, startrow=7)
            wb = load_workbook(new_filename)
            ws = wb.active
            ws['A1'] = 'UrlLink'
            ws['B1'] = 'ItemCode'
            ws['C1'] = 'ID'
            ws['D1'] = 'PW'

            ws['A2'] = 'https://www.redprinting.co.kr/ko/product/item/'+itemCode[:2]+'/'+itemCode
            ws['B2'] = itemCode
            ws['C2'] = 'x'
            ws['D2'] = 'x'

            wb.save(new_filename)


def uploadfile_ordernum_creating(df_item_config, df_item):
    itemList_card = ['BCSPDFT', 'BNSTDFT']
    itemList_sticker = ['STCUXXX', 'STTHCIC', 'STTHELP', 'STTHSQU']
    itemList_stationery = ['GSSBMTL', 'GSSBACM', 'GSSBSTP']
    options = ChromeOptions()
    options.add_argument('--blink-settings=imagesEnabled=false')
    driver = webdriver.Chrome(options=options)
    driver.implicitly_wait(3)
    totalList = len(df_item)

    # 상품코드 가져오기
    itemUrl = df_item_config.iloc[0,0]
    itemCode = df_item_config.iloc[0,1]
    userid = df_item_config.iloc[0,2]
    userpw = df_item_config.iloc[0,3]
    if userid == "x":
        userid = 'redprinting'
    if userpw == "x":
        userpw = 'redprinting#1234'
    time.sleep(1)
    driver.get(itemUrl)

    time.sleep(1)
    driver.find_element(By.CSS_SELECTOR, '#header > div.main-home-header-warp > div > div.header-right > ul > li.select-my > a').click()

    # 로그인화면 아이디 및 패스워드 입력
    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
    driver.find_element(By.ID, 'mb_id').click()
    driver.find_element(By.ID, 'mb_id').send_keys(userid)
    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
    driver.find_element(By.ID, 'mb_password').click()
    driver.find_element(By.ID, 'mb_password').send_keys(userpw)
    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
    driver.find_element(By.ID, 'btnLogin').click()
    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
    driver.execute_script('window.scrollTo(0, 200)')

    try:
        # 각 상품코드 별 어떤 형태인지 체크
        if itemCode in itemList_card:
            for i in range(totalList):
                paper = df_item.iloc[i, 1]
                wgtcod = df_item.iloc[i, 2]
                dosu = df_item.iloc[i, 3]
                size = df_item.iloc[i, 4]
                amount = df_item.iloc[i, 5]
                apcs1 = df_item.iloc[i, 6]
                apcs2 = df_item.iloc[i, 7]
                apcs3 = df_item.iloc[i, 8]
                apcs4 = df_item.iloc[i, 9]
                apcs5 = df_item.iloc[i, 10]
                sizesplit = size.split("*")
                wid_size = sizesplit[0]
                hei_size = sizesplit[1]
                driver.execute_script('window.scrollTo(0, 100)')
                time.sleep(0.5)
                paper_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "paperSelectBoxItText")) )
                if paper_text.text != paper:
                    # 상품 페이지 용지선택
                    time.sleep(0.5)
                    #paperSelectBoxItContainer
                    driver.find_element(By.ID, 'paperSelectBoxIt').click()
                    time.sleep(0.5)
                    driver.find_element(By.LINK_TEXT, paper).click()
                time.sleep(0.5)
                if wgtcod != "":
                    # 상품 페이지 G수 선택
                    if driver.find_element(By.ID, 'paper_sub_selectSelectBoxItText').text != str(wgtcod):
                        #paper_sub_selectSelectBoxItContainer
                        driver.find_element(By.ID, 'paper_sub_selectSelectBoxIt').click()
                        time.sleep(1)
                        driver.find_element(By.LINK_TEXT, str(wgtcod)).click()
                time.sleep(0.5)
                if dosu != "":
                    # 상품 페이지 인쇄도수 선택
                    docu_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "soduSelectBoxItText")) )
                    if docu_text.text != dosu:
                        #soduSelectBoxItContainer
                        driver.find_element(By.ID, 'soduSelectBoxIt').click()
                        WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                        driver.find_element(By.LINK_TEXT, dosu).click()
                time.sleep(0.5)
                if driver.find_element(By.ID, 'sizeSelectBoxItText').text != '사이즈직접입력':
                    # 사이즈 직접입력 선택 sizeSelectBoxItContainer
                    driver.find_element(By.ID, 'sizeSelectBoxIt').click()
                    time.sleep(1)
                    driver.find_element(By.LINK_TEXT, '사이즈직접입력').click()
                time.sleep(0.5)
                if driver.find_element(By.ID, 'CUT_WDT').get_attribute('value') != wid_size:
                    # 상품 페이지 사이즈 직접입력 사이즈 입력
                    driver.find_element(By.ID, 'CUT_WDT').click()
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'CUT_WDT').send_keys(wid_size)
                time.sleep(0.5)
                if driver.find_element(By.ID, 'CUT_HGH').get_attribute('value') != hei_size:
                    driver.find_element(By.ID, 'CUT_HGH').click()
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'CUT_HGH').send_keys(hei_size)
                time.sleep(0.5)
                driver.find_element(By.ID, 'WRK_HGH').click()
                time.sleep(0.5)
                if amount != "":
                    driver.execute_script("productOrder.check_PRN_CNT();")
                    number1_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "number1")) )
                    if number1_text.get_attribute('value') != str(amount):
                        time.sleep(0.5)
                        driver.find_element(By.ID, 'number1').click()
                        driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                        driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                        driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                        driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                        driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                        driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                        time.sleep(0.5)
                        driver.find_element(By.ID, 'number1').send_keys(str(amount))
                    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )

                time.sleep(0.5)
                apcs_nowstatus = driver.find_element(By.ID, 'opt_string').get_attribute('value')
                if apcs1 == "무광":
                    if 'COT_DFT' in apcs_nowstatus:
                        time.sleep(0.2)
                    else:
                        # 상품 페이지 코팅 선택
                        driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                    time.sleep(0.5)
                    # 상품 페이지 무광코팅 선택
                    driver.execute_script("productOrder.opt_select('COT_DFT','MA');")
                elif apcs1 == "유광":
                    if 'COT_DFT' in apcs_nowstatus:
                        time.sleep(0.2)
                    else:
                        # 상품 페이지 코팅 선택
                        driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                    time.sleep(0.5)
                    # 상품 페이지 유광코팅 선택
                    driver.execute_script("productOrder.opt_select('COT_DFT','GL');")
                else:
                    if 'COT_DFT' in apcs_nowstatus:
                        # 코팅 후가공 다시 선택 시 선택해제됨.
                        driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                    else:
                        time.sleep(0.2)

                if apcs2 == "x":
                    time.sleep(0.2)
                else:
                    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                    driver.execute_script("productOrder.opt_use_yn('HOL_DFT', '');")
                    aftpcs2 = apcs2.split('-')
                    apcs2_cnt = aftpcs2[0]
                    apcs2_val = aftpcs2[1]
                    driver.find_element(By.ID, 'HOL_DFT_IN1').click()
                    driver.find_element(By.ID, 'HOL_DFT_IN1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'HOL_DFT_IN1').send_keys(Keys.BACK_SPACE)
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'HOL_DFT_IN1').send_keys(str(apcs2_cnt))
                    # 라디오 버튼 네임값이 동일한 것 찾아서 apcs2 값 동일한 것으로 클릭 후 멈춤
                    afterPcs2List = driver.find_elements(By.NAME, 'HOL_DFT_MM')
                    for ap2l in afterPcs2List:
                        if ap2l.get_attribute('value') == apcs2_val:
                            ap2l.click()
                            break

                if apcs3 == 'x':
                    time.sleep(0.2)
                else:
                    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                    driver.execute_script("productOrder.opt_use_yn('ROU_DFT', '');")
                    aftpcs3 = apcs3.split('_')
                    apcs3_size = aftpcs3[0]
                    apcs3_location = str(aftpcs3[1])
                    afterPcs3List = driver.find_elements(By.NAME, 'ROU_DFT_MM')
                    for ap3l in afterPcs3List:
                        if ap3l.get_attribute('value') == apcs3_size:
                            ap3l.click()
                            break
                    apcs3location = driver.find_element(By.ID, 'ROU_DFT_ALL')

                    if apcs3_location == '1111' or apcs3_location == 'all':
                        time.sleep(0.2)
                    else:
                        if apcs3location.is_selected():
                            apcs3location.click()
                        else:
                            time.sleep(0.2)
                        if apcs3_location[0] == '1':
                            driver.execute_script("productOrder.rou_dft_select('DFXLT');")
                        if apcs3_location[1] == '1':
                            driver.execute_script("productOrder.rou_dft_select('DFXRT');")
                        if apcs3_location[2] == '1':
                            driver.execute_script("productOrder.rou_dft_select('DFXLB');")
                        if apcs3_location[3] == '1':
                            driver.execute_script("productOrder.rou_dft_select('DFXRB');")

                if apcs4 == 'x':
                    time.sleep(0.2)
                else:
                    WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    if apcs4 == '방풍커팅':
                        driver.execute_script("productOrder.opt_checked('CUT_ZUN', 'ZDWND');")

                if apcs5 == 'x':
                    time.sleep(0.2)
                else:
                    WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
                    if apcs5 == '은색아일렛':
                        driver.find_element(By.ID, 'ILT_DFT_IN1').click()
                    if apcs5 == '구리색아일렛':
                        driver.find_element(By.ID, 'ILT_DFT_IN2').click()

                # time.sleep(1)
                # driver.find_element(By.CSS_SELECTOR, 'body > div.page_order > div > aside > div.search_tool.search_tool_update > div.confirmbox-wrap > div.radio-group-big > span.radio-group-inner > label').click()
                # time.sleep(0.5)
                driver.execute_script('window.scrollTo(0, 300)')
                time.sleep(0.5)
                total_price = driver.find_element(By.ID, 'TOTAL_PRICE').text
                df_item.loc[i, 'Price'] = total_price
                print(i, "번째 TOTAL_PRICE : ", total_price)
                # print(df_item)
                # time.sleep(0.5)
                # driver.find_element(By.ID, 'direct_order_btn').click()
                # time.sleep(1)
                # al = Alert(driver)
                # al.accept()
                # time.sleep(0.5)
                # imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
                # df_item.loc[i, 'OrderCode'] = imsiordernum #주문관리코드생성후추가
                # print(i, "번째 TOTAL_PRICE : ", total_price, "pot_tmp_cod : ", imsiordernum)
                driver.execute_script('window.location.reload();')
                driver.execute_script('window.scrollTo(0, 100);')
                time.sleep(5)

        elif itemCode in itemList_sticker:
            for i in range(totalList):
                paper = df_item.iloc[i, 1]
                wgtcod = df_item.iloc[i, 2]
                dosu = df_item.iloc[i, 3]
                size = df_item.iloc[i, 4]
                amount = df_item.iloc[i, 5]
                apcs1 = df_item.iloc[i, 6]
                apcs2 = df_item.iloc[i, 7]
                apcs3 = df_item.iloc[i, 8]
                apcs4 = df_item.iloc[i, 9]
                apcs5 = df_item.iloc[i, 10]
                sizesplit = size.split("*")
                wid_size = sizesplit[0]
                hei_size = sizesplit[1]
                time.sleep(2)
                paper_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "paperSelectBoxItText")) )
                if paper_text.text != paper:
                    # 상품 페이지 용지선택
                    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                    # paperSelectBoxItContainer
                    driver.find_element(By.ID, 'paperSelectBoxIt').click()
                    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                    driver.find_element(By.LINK_TEXT, paper).click()

                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                if wgtcod != "":
                    # 상품 페이지 G수 선택
                    wgt_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "paper_sub_selectSelectBoxItText")) )
                    if wgt_text.text != str(wgtcod):
                        #paper_sub_selectSelectBoxItContainer
                        driver.find_element(By.ID, 'paper_sub_selectSelectBoxIt').click()
                        WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                        driver.find_element(By.LINK_TEXT, str(wgtcod)).click()

                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                if dosu != "":
                    # 상품 페이지 인쇄도수 선택
                    docu_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "soduSelectBoxItText")) )
                    if docu_text.text != dosu:
                        #soduSelectBoxItContainer
                        driver.find_element(By.ID, 'soduSelectBoxIt').click()
                        WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                        driver.find_element(By.LINK_TEXT, dosu).click()

                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                size_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "sizeSelectBoxItText")) )
                if size_text.text != '사이즈직접입력':
                    # 사이즈 직접입력 선택 sizeSelectBoxItContainer
                    driver.find_element(By.ID, 'sizeSelectBoxIt').click()
                    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                    driver.find_element(By.LINK_TEXT, '사이즈직접입력').click()

                CUT_WDT_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "CUT_WDT")) )
                if CUT_WDT_text.get_attribute('value') != wid_size:
                    # 상품 페이지 사이즈 직접입력 사이즈 입력
                    driver.find_element(By.ID, 'CUT_WDT').click()
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_WDT').send_keys(Keys.BACK_SPACE)
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'CUT_WDT').send_keys(wid_size)

                CUT_HGH_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "CUT_HGH")) )
                if CUT_HGH_text.get_attribute('value') != hei_size:
                    driver.find_element(By.ID, 'CUT_HGH').click()
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'CUT_HGH').send_keys(Keys.BACK_SPACE)
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'CUT_HGH').send_keys(hei_size)

                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                driver.find_element(By.XPATH, '//*[@id="WRK_HGH"]').click()
                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                driver.execute_script("productOrder.check_PRN_CNT();")
                number1_text = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "number1")) )
                if number1_text.get_attribute('value') != str(amount):
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'number1').click()
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.DELETE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    driver.find_element(By.ID, 'number1').send_keys(Keys.BACK_SPACE)
                    time.sleep(0.5)
                    driver.find_element(By.ID, 'number1').send_keys(str(amount))

                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                apcs_nowstatus = driver.find_element(By.ID, 'opt_string').get_attribute('value')
                if apcs1 == "무광":
                    if 'COT_DFT' in apcs_nowstatus:
                        time.sleep(0.2)
                    else:
                        # 상품 페이지 코팅 선택
                        driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                    driver.execute_script("productOrder.opt_select('COT_DFT','MA');")# 상품 페이지 무광코팅 선택
                elif apcs1 == "유광":
                    if 'COT_DFT' in apcs_nowstatus:
                        time.sleep(0.2)
                    else:
                        # 상품 페이지 코팅 선택
                        driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');")
                    WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                    driver.execute_script("productOrder.opt_select('COT_DFT','GL');")# 상품 페이지 유광코팅 선택
                else:
                    if 'COT_DFT' in apcs_nowstatus:
                        driver.execute_script("productOrder.opt_use_yn('COT_DFT', 'SID_S');") #코팅 후가공 다시 선택 시 선택해제됨.
                    else:
                        time.sleep(0.2)
                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )



                if apcs4 == "묶음재단":
                    driver.execute_script("productOrder.opt_checked('CUT_DFT', 'DFXXX');")  # 상품 페이지 묶음재단 선택
                elif apcs4 == "개별재단":
                    driver.execute_script("productOrder.opt_checked('CUT_DFT', 'DFITM');")  # 상품 페이지 개별재단 선택
                else:
                    time.sleep(0.2)
                WebDriverWait(driver, 10).until( EC.invisibility_of_element_located((By.ID, 'overlay')) )
                driver.find_element(By.XPATH, '//*[@id="WRK_HGH"]').click()
                total_price = WebDriverWait(driver, 5).until( EC.visibility_of_element_located((By.ID, "TOTAL_PRICE")) )
                tprice = total_price.text
                # total_price = driver.find_element(By.ID, 'TOTAL_PRICE').text
                df_item.loc[i, 'Price'] = tprice
                time.sleep(0.5)
                driver.find_element(By.ID, 'direct_order_btn').click()
                time.sleep(1)
                al = Alert(driver)
                al.accept()
                time.sleep(0.5)
                imsiordernum = driver.find_element(By.ID, 'pot_tmp_cod').get_attribute('value')
                df_item.loc[i, 'OrderCode'] = imsiordernum  # 주문관리코드생성후추가
                print(i, "번째 TOTAL_PRICE : ", tprice, "pot_tmp_cod : ", imsiordernum)
                time.sleep(5)
        else:
            print('error')

        nowtime=str(now.year)+'_'+str(now.month)+'_'+str(now.day)+'_'+str(now.hour)+'_'+str(now.minute)
        new_filename = REDATA_FOLDER+'/'+userid+'_'+itemCode+'_' + nowtime + '.xlsx'
        df_item.to_excel(new_filename, sheet_name=itemCode, index=False)
        time.sleep(5)
        driver.quit()
    except WebDriverException as e:
        print(f"WebDriver 오류 발생: {e}")
        driver.quit()
    finally:
        driver.quit()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
